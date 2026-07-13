"""
generar_resumen_stock_compra.py

Script mensual y autosuficiente: no depende de haber corrido antes ningún otro
script ni de que exista un consolidado.xlsx previo. Cada mes basta con:
  1. Reemplazar los archivos Sal_Art_*.xlsx (export de stock por bodega)
  2. Reemplazar Consumos_Historicos.xlsx (histórico de consumo por bodega)
  3. Ejecutar:  python generar_resumen_stock_compra.py

Genera dos archivos:
  - consolidado.xlsx       -> stock crudo consolidado de todas las bodegas (para auditar)
  - resumen_stock_BFC.xlsx -> saldos, proyecciones de consumo, necesidad de compra
                               y alertas de stock, centrado en la bodega BFC (droguería)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

carpeta = Path(__file__).resolve().parent

# ── Parámetros del negocio: tocar aquí si cambian las reglas ──────────────────
BODEGA_BFC              = "Sal_Art_BFC"     # bodega de droguería: la que compra/distribuye
BODEGAS_EXCLUIR_STOCK   = {"Sal_Art_ETHON"} # no son puntos reales de stock de la red
PATRON_BODEGA_TRANSITO  = "ransversal"      # "Unidad Técnica Transversal" (UTTO): tránsito interno, no consumo real
MESES_PROYECCION        = (3, 6)            # horizontes de proyección/compra que pide el negocio


# ════════════════════════════════════════════════════════════════════════════
# PASO 1: LEER Y CONSOLIDAR EL STOCK DE TODAS LAS BODEGAS (Sal_Art_*.xlsx)
# ════════════════════════════════════════════════════════════════════════════
def leer_archivo_stock(ruta):
    """Cada export Sal_Art_*.xlsx trae 2 filas de cabecera con celdas combinadas
    en una posición que puede variar de archivo en archivo. En vez de asumir un
    número de fila fijo, se detecta automáticamente la primera fila con al menos
    5 celdas no vacías y se arma el nombre de columna combinando esa fila con la
    siguiente (p.ej. "Stock" + "Máx." -> "Stock Máx.")."""
    wb = load_workbook(ruta, data_only=True)
    ws = wb.active

    fila_cab = None
    for i, fila in enumerate(ws.iter_rows(values_only=True), start=1):
        no_vacios = sum(1 for c in fila if c is not None and str(c).strip() != "")
        if no_vacios >= 5:
            fila_cab = i
            break
    if fila_cab is None:
        return pd.DataFrame()

    cab1 = [c.value for c in ws[fila_cab]]
    cab2 = [c.value for c in ws[fila_cab + 1]]

    columnas = []
    for a, b in zip(cab1, cab2):
        a = str(a).strip() if a is not None else ""
        b = str(b).strip() if b is not None else ""
        if a and b:
            columnas.append(f"{a} {b}")
        elif a:
            columnas.append(a)
        elif b:
            columnas.append(b)
        else:
            columnas.append("Sin nombre")

    filas = []
    for fila in ws.iter_rows(min_row=fila_cab + 2, values_only=True):
        if any(c is not None and str(c).strip() != "" for c in fila):
            filas.append(list(fila))

    df = pd.DataFrame(filas, columns=columnas)
    df = df.loc[:, ~df.columns.str.startswith("Sin nombre")]
    df = df.dropna(how="all")
    # En algunas bodegas "Código de Barras" viene vacío (dtype object) y en otras
    # trae valores numéricos (dtype float64); se fija como texto en todas para
    # que el dtype sea consistente antes de concatenar los archivos.
    if "Código de Barras" in df.columns:
        df["Código de Barras"] = df["Código de Barras"].astype("string")
    return df


archivos_stock = sorted(carpeta.glob("Sal_Art_*.xlsx"))
if not archivos_stock:
    raise FileNotFoundError(f"No se encontraron archivos Sal_Art_*.xlsx en {carpeta}")

trozos = []
for archivo in archivos_stock:
    origen = archivo.stem  # nombre del archivo sin extensión = identificador de la bodega
    df_bodega = leer_archivo_stock(archivo)
    if df_bodega.empty:
        print(f"  Aviso: {archivo.name} no aportó filas, se omite")
        continue
    df_bodega.insert(0, "origen", origen)
    trozos.append(df_bodega)

cons = pd.concat(trozos, ignore_index=True, sort=False)

# Normalizar tipos numéricos (vienen como texto/objeto tras leer con openpyxl)
cons["Código"]       = pd.to_numeric(cons["Código"], errors="coerce")
cons["Stock Actual"] = pd.to_numeric(cons["Stock Actual"], errors="coerce").fillna(0)
cons["Stock Mínimo"] = pd.to_numeric(cons["Stock Mínimo"], errors="coerce").fillna(0)
cons["Stock Máx."]   = pd.to_numeric(cons["Stock Máx."], errors="coerce").fillna(0)
cons = cons.dropna(subset=["Código"])
cons["Código"] = cons["Código"].astype(int)

print(f"Stock consolidado: {len(cons)} filas de {len(archivos_stock)} bodegas "
      f"({cons['Código'].nunique()} productos únicos)")

# Se guarda el consolidado crudo para poder auditar el cruce si algo no cuadra
cons.to_excel(carpeta / "consolidado.xlsx", index=False)


# ════════════════════════════════════════════════════════════════════════════
# PASO 2: CALCULAR LOS SALDOS QUE PIDE EL RESUMEN, A PARTIR DEL STOCK CONSOLIDADO
# ════════════════════════════════════════════════════════════════════════════
# 2a. Saldo, mínimo y máximo de la bodega BFC (droguería) por producto
bfc = cons[cons["origen"] == BODEGA_BFC].copy()
if bfc.empty:
    raise ValueError(f"No se encontró la bodega '{BODEGA_BFC}' entre los Sal_Art_*.xlsx")

bfc["Ult. Precio Ingresado"] = pd.to_numeric(bfc["Ult. Precio Ingresado"], errors="coerce")
saldo_bfc = bfc.groupby("Código", as_index=False).agg(
    **{"Saldo BFC":  ("Stock Actual",          "sum"),
       "Mínimo BFC": ("Stock Mínimo",          "sum"),
       "Máximo BFC": ("Stock Máx.",            "sum"),
       "Ult. Precio": ("Ult. Precio Ingresado", "max")}
)

# 2b. Total de saldos de toda la red (todas las bodegas, excepto las que no son
#     puntos reales de stock, p.ej. ETHON)
red = cons[~cons["origen"].isin(BODEGAS_EXCLUIR_STOCK)]
total_saldos = (red.groupby("Código", as_index=False)["Stock Actual"].sum()
                 .rename(columns={"Stock Actual": "Total Saldos"}))

# 2c. Total de máximos de la red EXCLUYENDO BFC (para dimensionar cuánto stock
#     máximo sostiene el resto de la red, sin contar la droguería)
resto = red[red["origen"] != BODEGA_BFC]
total_max_resto = (resto.groupby("Código", as_index=False)["Stock Máx."].sum()
                    .rename(columns={"Stock Máx.": "Total Máximos (excepto BFC)"}))

# 2d. Nombre y proveedor "canónicos" de cada producto (primera aparición en el stock)
nombres = (cons.drop_duplicates("Código")[["Código", "Nombre Artículo", "Proveedor"]]
           .rename(columns={"Nombre Artículo": "Medicamento"}))


# ════════════════════════════════════════════════════════════════════════════
# PASO 3: LEER EL CONSUMO HISTÓRICO Y CALCULAR EL CONSUMO MENSUAL PROMEDIO
# ════════════════════════════════════════════════════════════════════════════
# Nota: las columnas "Código"/"Artículo" de este archivo llegan con la
# codificación dañada de origen (problema del sistema que lo exporta), por eso
# se identifican por posición (columna 0 y 1) en vez de por nombre literal.
ch = pd.read_excel(carpeta / "Consumos_Historicos.xlsx", sheet_name="Hoja1")
col_codigo, col_articulo = ch.columns[0], ch.columns[1]
ch = ch.rename(columns={col_codigo: "Codigo", col_articulo: "Articulo"})
ch["Codigo"] = pd.to_numeric(ch["Codigo"], errors="coerce")
ch["Fecha"]  = pd.to_datetime(ch["Fecha"])
ch = ch.dropna(subset=["Codigo"])
ch["Codigo"] = ch["Codigo"].astype(int)

# Columnas de bodega = todas menos Codigo/Articulo/TOTAL/Fecha.
# Se excluye la bodega de tránsito interno (UTTO/"Unidad Técnica Transversal"):
# no es consumo real, es solo traspaso entre bodegas.
cols_bodega  = [c for c in ch.columns if c not in ("Codigo", "Articulo", "TOTAL", "Fecha")]
col_transito = [c for c in cols_bodega if PATRON_BODEGA_TRANSITO in c and "Sur" not in c]
cols_consumo = [c for c in cols_bodega if c not in col_transito]

# Consumo mensual promedio de los últimos 12 meses CON DATOS. Es dinámico: cada
# vez que se reemplaza el archivo, se recalcula solo sobre los 12 meses más
# recientes según la fecha máxima que traiga el archivo nuevo.
fecha_max    = ch["Fecha"].max()
fecha_inicio = fecha_max - pd.DateOffset(months=11)
reciente = ch[ch["Fecha"] >= fecha_inicio].copy()
reciente[cols_consumo] = reciente[cols_consumo].apply(pd.to_numeric, errors="coerce").fillna(0)
reciente["consumo_red"] = reciente[cols_consumo].sum(axis=1)  # consumo de toda la red por mes

# Se divide por la cantidad real de meses con registro de cada producto (no
# todos los productos tienen historial en los 12 meses completos)
meses_por_codigo = reciente.groupby("Codigo")["Fecha"].nunique()
consumo = reciente.groupby("Codigo", as_index=False)["consumo_red"].sum()
consumo["meses"]           = consumo["Codigo"].map(meses_por_codigo).fillna(1)
consumo["Consumo Mensual"] = (consumo["consumo_red"] / consumo["meses"]).round(2)
consumo = consumo.rename(columns={"Codigo": "Código"})[["Código", "Consumo Mensual"]]

# Universo de productos del resumen: el arsenal de droguería (los que tienen
# historial de consumo), igual que en el proceso de generación de pedidos
arsenal = (ch[["Codigo", "Articulo"]].drop_duplicates("Codigo")
           .rename(columns={"Codigo": "Código"}))

print(f"Consumo calculado sobre {fecha_inicio.date()} a {fecha_max.date()} "
      f"({len(arsenal)} productos, bodega de tránsito excluida: {col_transito})")


# ════════════════════════════════════════════════════════════════════════════
# PASO 4: CRUZAR STOCK + CONSUMO EN UNA SOLA TABLA
# ════════════════════════════════════════════════════════════════════════════
df = arsenal.merge(saldo_bfc,       on="Código", how="left")
df = df.merge(total_saldos,         on="Código", how="left")
df = df.merge(total_max_resto,      on="Código", how="left")
df = df.merge(nombres,              on="Código", how="left")
df = df.merge(consumo,              on="Código", how="left")

# Productos del arsenal sin stock/consumo cruzado quedan en 0, no en NaN
for col in ["Saldo BFC", "Mínimo BFC", "Máximo BFC", "Total Saldos",
            "Total Máximos (excepto BFC)", "Consumo Mensual", "Ult. Precio"]:
    df[col] = df[col].fillna(0)
df["Proveedor"]   = df["Proveedor"].fillna("")
df["Medicamento"] = df["Medicamento"].fillna(df["Articulo"])  # si no hay nombre de stock, usa el del histórico

# Valorización de stock al último precio ingresado en BFC (mismo criterio de
# precio que usaba generar_pedido.py)
df["Valor Stock BFC"]   = (df["Saldo BFC"]    * df["Ult. Precio"]).round(0).astype(int)
df["Valor Stock Total"] = (df["Total Saldos"] * df["Ult. Precio"]).round(0).astype(int)


# ════════════════════════════════════════════════════════════════════════════
# PASO 5: PROYECCIÓN DE CONSUMO, NECESIDAD DE COMPRA Y ALERTAS DE STOCK
# ════════════════════════════════════════════════════════════════════════════
# Proyección de consumo a N meses = cuánto se espera consumir en ese horizonte
# Necesidad de compra a N meses  = lo que falta comprar para cubrir esa proyección
#                                   con el stock total actual de la red (sin bajar de 0)
for m in MESES_PROYECCION:
    df[f"Proyección Consumo {m}M"] = (df["Consumo Mensual"] * m).round(0).astype(int)
    df[f"Necesidad Compra {m}M"]   = (
        (df[f"Proyección Consumo {m}M"] - df["Total Saldos"]).clip(lower=0).round(0).astype(int)
    )

# Alerta de stock según meses de cobertura de la red completa
# (Total Saldos / Consumo Mensual): cuántos meses dura el stock actual al ritmo
# de consumo promedio.
consumo_nz      = df["Consumo Mensual"].replace(0, np.nan)
meses_cobertura = df["Total Saldos"] / consumo_nz

def alerta(m):
    if pd.isna(m): return "SIN CONSUMO"  # producto sin consumo registrado: no se puede proyectar
    if m < 1:      return "URGENTE"      # queda menos de 1 mes de stock
    if m <= 2:     return "NORMAL"       # entre 1 y 2 meses de stock
    return "OK"                          # más de 2 meses de stock

df["Alerta Stock"] = meses_cobertura.map(alerta)

# Se ordena mostrando primero lo más urgente y, dentro de cada nivel, lo que
# requiere comprar más cantidad a 3 meses
orden_a = {"URGENTE": 0, "NORMAL": 1, "OK": 2, "SIN CONSUMO": 3}
df["_ord"] = df["Alerta Stock"].map(orden_a)
df = df.sort_values(["_ord", "Necesidad Compra 3M"], ascending=[True, False]).drop(columns="_ord")

cols_salida = [
    "Código", "Medicamento", "Proveedor",
    "Saldo BFC", "Mínimo BFC", "Máximo BFC",
    "Total Saldos", "Total Máximos (excepto BFC)",
    "Ult. Precio", "Valor Stock BFC", "Valor Stock Total",
    "Consumo Mensual",
    "Proyección Consumo 3M", "Proyección Consumo 6M",
    "Necesidad Compra 3M", "Necesidad Compra 6M",
    "Alerta Stock",
]
salida = df[cols_salida].reset_index(drop=True)

print(f"\nProductos en resumen: {len(salida)}")
print(f"Alertas:\n{salida['Alerta Stock'].value_counts().to_string()}")


# ════════════════════════════════════════════════════════════════════════════
# PASO 6: GUARDAR EL RESUMEN EN EXCEL, CON FORMATO Y COLORES POR ALERTA
# ════════════════════════════════════════════════════════════════════════════
fill_urg  = PatternFill(fill_type="solid", fgColor="FF4444")  # rojo    -> URGENTE
fill_nor  = PatternFill(fill_type="solid", fgColor="FFD966")  # amarillo-> NORMAL
fill_ok   = PatternFill(fill_type="solid", fgColor="70AD47")  # verde   -> OK
fill_sc   = PatternFill(fill_type="solid", fgColor="BFBFBF")  # gris    -> SIN CONSUMO
fill_hdr  = PatternFill(fill_type="solid", fgColor="1F4E79")
font_hdr  = Font(bold=True, color="FFFFFF", size=10)
font_bold = Font(bold=True)
align_c   = Alignment(horizontal="center", vertical="center")

anchos = [10, 50, 28, 14, 14, 14, 16, 24, 14, 18, 18, 18, 20, 20, 18, 18, 14]

def escribir_excel(path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        salida.to_excel(writer, index=False, sheet_name="Resumen Stock BFC")
        ws = writer.sheets["Resumen Stock BFC"]

        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        for cell in ws[1]:
            cell.fill      = fill_hdr
            cell.font      = font_hdr
            cell.alignment = align_c
        ws.row_dimensions[1].height = 30

        colores = {"URGENTE": fill_urg, "NORMAL": fill_nor, "OK": fill_ok, "SIN CONSUMO": fill_sc}
        col_alerta = len(cols_salida)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            alerta_val = row[col_alerta - 1].value
            fill       = colores.get(alerta_val)
            if fill:
                row[col_alerta - 1].fill = fill
                row[col_alerta - 1].font = font_bold
            for cell in row[2:]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row[1].alignment = Alignment(horizontal="left", vertical="center")

        ws.freeze_panes = "A2"


out_primary = carpeta / "resumen_stock_BFC.xlsx"
out_backup  = carpeta / "resumen_stock_BFC_backup.xlsx"

try:
    escribir_excel(out_primary)
    print(f"\nArchivo guardado: {out_primary}")
except PermissionError:
    # Si el archivo está abierto en Excel no se puede sobrescribir: se guarda
    # una copia de respaldo para no perder el resultado del cálculo
    escribir_excel(out_backup)
    print(f"\n{out_primary.name} está abierto en Excel y no se pudo sobrescribir.")
    print(f"Se guardó como: {out_backup}")
    print("Ciérralo y vuelve a ejecutar el script, o renombra el archivo de respaldo manualmente.")
