"""
generar_dashboard_completo.py

Pipeline completo en un solo script: de los exports crudos de stock por bodega
al dashboard HTML, sin pasos intermedios manuales. Cada mes basta con:
  1. Reemplazar los archivos Sal_Art_*.xlsx (export de stock por bodega)
  2. Reemplazar Consumos_Historicos.xlsx (histórico de consumo por bodega)
  3. Ejecutar:  python generar_dashboard_completo.py

Genera:
  - consolidado.xlsx        -> stock crudo consolidado de todas las bodegas (para auditar)
  - resumen_stock_BFC.xlsx  -> saldos, valorización, proyecciones y alertas (para auditar/imprimir)
  - dashboard_farmacia.html -> dashboard interactivo, 100% autocontenido (sin internet)

Es la unión de generar_resumen_stock_compra.py + generar_dashboard.py: hace el
mismo cálculo pero en una sola pasada, leyendo Consumos_Historicos.xlsx una
sola vez, y sin depender de que exista ya un resumen_stock_BFC.xlsx previo con
el esquema de columnas correcto.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import re
import json
import pickle
import hashlib
import pandas as pd
import numpy as np
from plotly.offline import get_plotlyjs
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

import pronostico_demanda

carpeta = Path(__file__).resolve().parent

TOTAL_PASOS = 10


def _log_paso(n, mensaje):
    """Indicador de avance del pipeline (10 pasos, de la lectura de los
    Excel de origen a la escritura final del dashboard HTML)."""
    print(f"\n[Paso {n}/{TOTAL_PASOS}] {mensaje}")


# ── Parámetros del negocio: tocar aquí si cambian las reglas ──────────────────
BODEGA_BFC              = "Sal_Art_BFC"     # bodega de droguería: la que compra/distribuye
BODEGAS_EXCLUIR_STOCK   = {"Sal_Art_ETHON"} # no son puntos reales de stock de la red
PATRON_BODEGA_TRANSITO  = "ransversal"      # "Unidad Técnica Transversal" (UTTO): tránsito interno, no consumo real
MESES_PROYECCION        = (3, 6)            # horizontes de proyección/compra que pide el negocio

RUTA_CONSUMOS = carpeta / "Consumos_Historicos.xlsx"
RUTA_RESUMEN  = carpeta / "resumen_stock_BFC.xlsx"
RUTA_DASHBOARD = carpeta / "dashboard_farmacia.html"


def nombre_centro(col):
    """Limpia el nombre largo de bodega para usarlo como etiqueta
    (p.ej. 'Bodega Farmacia Dr. Abel zapata' -> 'Dr. Abel Zapata')."""
    n = re.sub(r"^bodega\s+(farmacia|farnacia)?\s*", "", col, flags=re.IGNORECASE).strip()
    if n.isupper():
        n = n.title()
    return n or col


# ════════════════════════════════════════════════════════════════════════════
# PASO 1: LEER Y CONSOLIDAR EL STOCK DE TODAS LAS BODEGAS (Sal_Art_*.xlsx)
# ════════════════════════════════════════════════════════════════════════════
def leer_archivo_stock(ruta):
    """Cada export Sal_Art_*.xlsx trae 2 filas de cabecera con celdas combinadas
    en una posición que puede variar de archivo en archivo. En vez de asumir un
    número de fila fijo, se detecta automáticamente la primera fila con al menos
    5 celdas no vacías y se arma el nombre de columna combinando esa fila con la
    siguiente (p.ej. "Stock" + "Máx." -> "Stock Máx.")."""
    wb = load_workbook(ruta, data_only=True)
    ws = wb.active

    fila_cab = None
    for i, fila in enumerate(ws.iter_rows(values_only=True), start=1):
        no_vacios = sum(1 for c in fila if c is not None and str(c).strip() != "")
        if no_vacios >= 5:
            fila_cab = i
            break
    if fila_cab is None:
        return pd.DataFrame()

    cab1 = [c.value for c in ws[fila_cab]]
    cab2 = [c.value for c in ws[fila_cab + 1]]

    columnas = []
    for a, b in zip(cab1, cab2):
        a = str(a).strip() if a is not None else ""
        b = str(b).strip() if b is not None else ""
        if a and b:
            columnas.append(f"{a} {b}")
        elif a:
            columnas.append(a)
        elif b:
            columnas.append(b)
        else:
            columnas.append("Sin nombre")

    filas = []
    for fila in ws.iter_rows(min_row=fila_cab + 2, values_only=True):
        if any(c is not None and str(c).strip() != "" for c in fila):
            filas.append(list(fila))

    df = pd.DataFrame(filas, columns=columnas)
    df = df.loc[:, ~df.columns.str.startswith("Sin nombre")]
    df = df.dropna(how="all")
    # En algunas bodegas "Código de Barras" viene vacío (dtype object) y en otras
    # trae valores numéricos (dtype float64); se fija como texto en todas para
    # que el dtype sea consistente antes de concatenar los archivos.
    if "Código de Barras" in df.columns:
        df["Código de Barras"] = df["Código de Barras"].astype("string")
    return df


_log_paso(1, "Leyendo y consolidando el stock de todas las bodegas (Sal_Art_*.xlsx)...")
archivos_stock = sorted(carpeta.glob("Sal_Art_*.xlsx"))
if not archivos_stock:
    raise FileNotFoundError(f"No se encontraron archivos Sal_Art_*.xlsx en {carpeta}")

trozos = []
for archivo in archivos_stock:
    origen = archivo.stem  # nombre del archivo sin extensión = identificador de la bodega
    df_bodega = leer_archivo_stock(archivo)
    if df_bodega.empty:
        print(f"  Aviso: {archivo.name} no aportó filas, se omite")
        continue
    df_bodega.insert(0, "origen", origen)
    trozos.append(df_bodega)

cons = pd.concat(trozos, ignore_index=True, sort=False)

# Normalizar tipos numéricos (vienen como texto/objeto tras leer con openpyxl)
cons["Código"]       = pd.to_numeric(cons["Código"], errors="coerce")
cons["Stock Actual"] = pd.to_numeric(cons["Stock Actual"], errors="coerce").fillna(0)
cons["Stock Mínimo"] = pd.to_numeric(cons["Stock Mínimo"], errors="coerce").fillna(0)
cons["Stock Máx."]   = pd.to_numeric(cons["Stock Máx."], errors="coerce").fillna(0)
cons = cons.dropna(subset=["Código"])
cons["Código"] = cons["Código"].astype(int)

print(f"Stock consolidado: {len(cons)} filas de {len(archivos_stock)} bodegas "
      f"({cons['Código'].nunique()} productos únicos)")

# Se guarda el consolidado crudo para poder auditar el cruce si algo no cuadra
cons.to_excel(carpeta / "consolidado.xlsx", index=False)


# ════════════════════════════════════════════════════════════════════════════
# PASO 2: CALCULAR LOS SALDOS QUE PIDE EL RESUMEN, A PARTIR DEL STOCK CONSOLIDADO
# ════════════════════════════════════════════════════════════════════════════
_log_paso(2, "Calculando saldos de stock (BFC y red completa)...")
bfc = cons[cons["origen"] == BODEGA_BFC].copy()
if bfc.empty:
    raise ValueError(f"No se encontró la bodega '{BODEGA_BFC}' entre los Sal_Art_*.xlsx")

bfc["Ult. Precio Ingresado"] = pd.to_numeric(bfc["Ult. Precio Ingresado"], errors="coerce")
saldo_bfc = bfc.groupby("Código", as_index=False).agg(
    **{"Saldo BFC":  ("Stock Actual",          "sum"),
       "Mínimo BFC": ("Stock Mínimo",          "sum"),
       "Máximo BFC": ("Stock Máx.",            "sum"),
       "Ult. Precio": ("Ult. Precio Ingresado", "max")}
)

red = cons[~cons["origen"].isin(BODEGAS_EXCLUIR_STOCK)]
total_saldos = (red.groupby("Código", as_index=False)["Stock Actual"].sum()
                 .rename(columns={"Stock Actual": "Total Saldos"}))

resto = red[red["origen"] != BODEGA_BFC]
total_max_resto = (resto.groupby("Código", as_index=False)["Stock Máx."].sum()
                    .rename(columns={"Stock Máx.": "Total Máximos (excepto BFC)"}))

# Saldo por establecimiento (una fila por bodega de la red): se usa tanto
# para el total de la red (panel resumen sin producto seleccionado) como
# para el desglose por producto (panel filtrado al medicamento elegido).
precio_por_codigo = dict(zip(saldo_bfc["Código"], saldo_bfc["Ult. Precio"]))
red_est = red.copy()
red_est["centro"]     = red_est["origen"].str.replace("Sal_Art_", "", regex=False)
red_est["_precio"]    = red_est["Código"].map(precio_por_codigo).fillna(0)
red_est["_valor"]     = red_est["Stock Actual"] * red_est["_precio"]
red_est["_con_stock"] = (red_est["Stock Actual"] > 0).astype(int)
saldo_por_establecimiento = (
    red_est.groupby("centro", as_index=False)
    .agg(unidades=("Stock Actual", "sum"),
         productos=("_con_stock", "sum"),
         valor=("_valor", "sum"))
    .sort_values("valor", ascending=False)
)

# Desglose de saldo por producto x establecimiento (para el panel cuando hay
# un medicamento seleccionado): una matriz Código -> [unidades por centro],
# alineada con `centros_stock` en el mismo orden para todos los productos.
centros_stock = sorted(set(red_est["centro"]))
_pivot_stock = (
    red_est.groupby(["Código", "centro"], as_index=False)["Stock Actual"].sum()
    .pivot(index="Código", columns="centro", values="Stock Actual")
    .reindex(columns=centros_stock, fill_value=0)
    .fillna(0)
)
saldo_centros_por_codigo = {
    int(cod): [float(v) for v in fila] for cod, fila in zip(_pivot_stock.index, _pivot_stock.values)
}

nombres = (cons.drop_duplicates("Código")[["Código", "Nombre Artículo", "Proveedor"]]
           .rename(columns={"Nombre Artículo": "Medicamento"}))

# Categoría/programa: columna presente en los archivos de stock pero vacía en
# casi todos los productos (limitación de origen, documentada en Metodología).
# Se guarda aparte (no se agrega a `salida`/Excel) solo para poder ofrecer el
# filtro de categoría en el dashboard sin tocar el esquema del resumen actual.
if "Categoría" in cons.columns:
    categoria_por_codigo = dict(zip(
        cons.drop_duplicates("Código")["Código"], cons.drop_duplicates("Código")["Categoría"]
    ))
else:
    categoria_por_codigo = {}


def _categoria_de(codigo):
    val = categoria_por_codigo.get(int(codigo))
    return str(val).strip() if pd.notna(val) and str(val).strip() else "Sin categoría"


# ════════════════════════════════════════════════════════════════════════════
# PASO 3: LEER EL CONSUMO HISTÓRICO — se usa tanto para el promedio mensual
# (resumen/proyecciones) como para la serie mes a mes del dashboard, en una
# sola lectura del archivo.
# ════════════════════════════════════════════════════════════════════════════
_log_paso(3, "Leyendo el consumo histórico (Consumos_Historicos.xlsx)...")
if not RUTA_CONSUMOS.exists():
    raise FileNotFoundError(f"No se encontró {RUTA_CONSUMOS.name} en {carpeta}")

# Nota: las columnas "Código"/"Artículo" de este archivo llegan con la
# codificación dañada de origen (problema del sistema que lo exporta), por eso
# se identifican por posición (columna 0 y 1) en vez de por nombre literal.
ch = pd.read_excel(RUTA_CONSUMOS, sheet_name="Hoja1")
col_codigo, col_articulo = ch.columns[0], ch.columns[1]
ch = ch.rename(columns={col_codigo: "Codigo", col_articulo: "Articulo"})
ch["Codigo"] = pd.to_numeric(ch["Codigo"], errors="coerce")
ch["Fecha"]  = pd.to_datetime(ch["Fecha"])
ch = ch.dropna(subset=["Codigo"])
ch["Codigo"] = ch["Codigo"].astype(int)
ch["Mes"]    = ch["Fecha"].dt.strftime("%Y-%m")

# Columnas de bodega = todas menos Codigo/Articulo/TOTAL/Fecha/Mes. Se dejan
# todas (incluida la de tránsito UTTO) para el desglose por centro del
# dashboard; el promedio de consumo mensual usado en proyecciones/compra la
# excluye más abajo, porque es solo traspaso interno, no consumo real.
cols_centro = [c for c in ch.columns if c not in ("Codigo", "Articulo", "TOTAL", "Fecha", "Mes")]
ch[cols_centro] = ch[cols_centro].apply(pd.to_numeric, errors="coerce").fillna(0)
col_transito = [c for c in cols_centro if PATRON_BODEGA_TRANSITO in c and "Sur" not in c]
cols_consumo = [c for c in cols_centro if c not in col_transito]

# Consumo mensual promedio de los últimos 12 meses CON DATOS. Es dinámico: cada
# vez que se reemplaza el archivo, se recalcula solo sobre los 12 meses más
# recientes según la fecha máxima que traiga el archivo nuevo.
fecha_max    = ch["Fecha"].max()
fecha_inicio = fecha_max - pd.DateOffset(months=11)
reciente = ch[ch["Fecha"] >= fecha_inicio].copy()
reciente["consumo_red"] = reciente[cols_consumo].sum(axis=1)  # consumo de toda la red por mes

# Se divide por la cantidad real de meses con registro de cada producto (no
# todos los productos tienen historial en los 12 meses completos)
meses_por_codigo = reciente.groupby("Codigo")["Fecha"].nunique()
consumo = reciente.groupby("Codigo", as_index=False)["consumo_red"].sum()
consumo["meses"]           = consumo["Codigo"].map(meses_por_codigo).fillna(1)
consumo["Consumo Mensual"] = (consumo["consumo_red"] / consumo["meses"]).round(2)
consumo = consumo.rename(columns={"Codigo": "Código"})[["Código", "Consumo Mensual"]]

# Universo de productos del resumen: el arsenal de droguería (los que tienen
# historial de consumo), igual que en el proceso de generación de pedidos
arsenal = (ch[["Codigo", "Articulo"]].drop_duplicates("Codigo")
           .rename(columns={"Codigo": "Código"}))

print(f"Consumo calculado sobre {fecha_inicio.date()} a {fecha_max.date()} "
      f"({len(arsenal)} productos, bodega de tránsito excluida: {col_transito})")


# ════════════════════════════════════════════════════════════════════════════
# PASO 4: CRUZAR STOCK + CONSUMO EN UNA SOLA TABLA
# ════════════════════════════════════════════════════════════════════════════
_log_paso(4, "Cruzando stock y consumo en una sola tabla...")
df = arsenal.merge(saldo_bfc,       on="Código", how="left")
df = df.merge(total_saldos,         on="Código", how="left")
df = df.merge(total_max_resto,      on="Código", how="left")
df = df.merge(nombres,              on="Código", how="left")
df = df.merge(consumo,              on="Código", how="left")

# Productos del arsenal sin stock/consumo cruzado quedan en 0, no en NaN
for col in ["Saldo BFC", "Mínimo BFC", "Máximo BFC", "Total Saldos",
            "Total Máximos (excepto BFC)", "Consumo Mensual", "Ult. Precio"]:
    df[col] = df[col].fillna(0)
df["Proveedor"]   = df["Proveedor"].fillna("")
df["Medicamento"] = df["Medicamento"].fillna(df["Articulo"])  # si no hay nombre de stock, usa el del histórico

# Valorización de stock al último precio ingresado en BFC
df["Valor Stock BFC"]   = (df["Saldo BFC"]    * df["Ult. Precio"]).round(0).astype(int)
df["Valor Stock Total"] = (df["Total Saldos"] * df["Ult. Precio"]).round(0).astype(int)


# ════════════════════════════════════════════════════════════════════════════
# PASO 5: PROYECCIÓN DE CONSUMO, NECESIDAD DE COMPRA Y ALERTAS DE STOCK
# ════════════════════════════════════════════════════════════════════════════
_log_paso(5, "Calculando proyecciones, necesidad de compra y alertas de stock...")
for m in MESES_PROYECCION:
    df[f"Proyección Consumo {m}M"] = (df["Consumo Mensual"] * m).round(0).astype(int)
    df[f"Necesidad Compra {m}M"]   = (
        (df[f"Proyección Consumo {m}M"] - df["Total Saldos"]).clip(lower=0).round(0).astype(int)
    )

consumo_nz      = df["Consumo Mensual"].replace(0, np.nan)
meses_cobertura = df["Total Saldos"] / consumo_nz

def alerta(m):
    if pd.isna(m): return "SIN CONSUMO"
    if m < 1:      return "URGENTE"
    if m <= 2:     return "NORMAL"
    return "OK"

df["Alerta Stock"] = meses_cobertura.map(alerta)

orden_a = {"URGENTE": 0, "NORMAL": 1, "OK": 2, "SIN CONSUMO": 3}
df["_ord"] = df["Alerta Stock"].map(orden_a)
df = df.sort_values(["_ord", "Necesidad Compra 3M"], ascending=[True, False]).drop(columns="_ord")

cols_salida = [
    "Código", "Medicamento", "Proveedor",
    "Saldo BFC", "Mínimo BFC", "Máximo BFC",
    "Total Saldos", "Total Máximos (excepto BFC)",
    "Ult. Precio", "Valor Stock BFC", "Valor Stock Total",
    "Consumo Mensual",
    "Proyección Consumo 3M", "Proyección Consumo 6M",
    "Necesidad Compra 3M", "Necesidad Compra 6M",
    "Alerta Stock",
]
salida = df[cols_salida].reset_index(drop=True)

print(f"\nProductos en resumen: {len(salida)}")
print(f"Alertas:\n{salida['Alerta Stock'].value_counts().to_string()}")


# ════════════════════════════════════════════════════════════════════════════
# PASO 6: GUARDAR EL RESUMEN EN EXCEL, CON FORMATO Y COLORES POR ALERTA
# (queda como respaldo/auditoría; el dashboard ya no depende de releerlo)
# ════════════════════════════════════════════════════════════════════════════
_log_paso(6, "Guardando el resumen en Excel (resumen_stock_BFC.xlsx)...")
fill_urg  = PatternFill(fill_type="solid", fgColor="FF4444")
fill_nor  = PatternFill(fill_type="solid", fgColor="FFD966")
fill_ok   = PatternFill(fill_type="solid", fgColor="70AD47")
fill_sc   = PatternFill(fill_type="solid", fgColor="BFBFBF")
fill_hdr  = PatternFill(fill_type="solid", fgColor="1F4E79")
font_hdr  = Font(bold=True, color="FFFFFF", size=10)
font_bold = Font(bold=True)
align_c   = Alignment(horizontal="center", vertical="center")

anchos = [10, 50, 28, 14, 14, 14, 16, 24, 14, 18, 18, 18, 20, 20, 18, 18, 14]

def escribir_excel(path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        salida.to_excel(writer, index=False, sheet_name="Resumen Stock BFC")
        ws = writer.sheets["Resumen Stock BFC"]

        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        for cell in ws[1]:
            cell.fill      = fill_hdr
            cell.font      = font_hdr
            cell.alignment = align_c
        ws.row_dimensions[1].height = 30

        colores = {"URGENTE": fill_urg, "NORMAL": fill_nor, "OK": fill_ok, "SIN CONSUMO": fill_sc}
        col_alerta = len(cols_salida)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            alerta_val = row[col_alerta - 1].value
            fill       = colores.get(alerta_val)
            if fill:
                row[col_alerta - 1].fill = fill
                row[col_alerta - 1].font = font_bold
            for cell in row[2:]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row[1].alignment = Alignment(horizontal="left", vertical="center")

        ws.freeze_panes = "A2"


out_backup = carpeta / "resumen_stock_BFC_backup.xlsx"
try:
    escribir_excel(RUTA_RESUMEN)
    print(f"\nArchivo guardado: {RUTA_RESUMEN}")
except PermissionError:
    escribir_excel(out_backup)
    print(f"\n{RUTA_RESUMEN.name} está abierto en Excel y no se pudo sobrescribir.")
    print(f"Se guardó como: {out_backup}")


# ════════════════════════════════════════════════════════════════════════════
# PASO 7: ARMAR LOS DATOS PARA EL BUSCADOR / TABLA DEL DASHBOARD (uno por producto)
# ════════════════════════════════════════════════════════════════════════════
_log_paso(7, "Armando los datos para el buscador y la tabla del dashboard...")
productos = []
for _, r in salida.iterrows():
    productos.append({
        "codigo":         int(r["Código"]),
        "medicamento":    str(r["Medicamento"]),
        "proveedor":      str(r["Proveedor"]) if pd.notna(r["Proveedor"]) else "",
        "saldoBfc":       float(r["Saldo BFC"]),
        "minimoBfc":      float(r["Mínimo BFC"]),
        "maximoBfc":      float(r["Máximo BFC"]),
        "totalSaldos":    float(r["Total Saldos"]),
        "totalMaxResto":  float(r["Total Máximos (excepto BFC)"]),
        "ultPrecio":      float(r["Ult. Precio"]),
        "valorStockBfc":  float(r["Valor Stock BFC"]),
        "valorStockTotal": float(r["Valor Stock Total"]),
        "consumoMensual": float(r["Consumo Mensual"]),
        "proy3":          float(r["Proyección Consumo 3M"]),
        "proy6":          float(r["Proyección Consumo 6M"]),
        "compra3":        float(r["Necesidad Compra 3M"]),
        "compra6":        float(r["Necesidad Compra 6M"]),
        "alerta":         str(r["Alerta Stock"]),
        "categoria":      _categoria_de(r["Código"]),
        "saldoCentros":   saldo_centros_por_codigo.get(int(r["Código"]), [0.0] * len(centros_stock)),
    })

codigos_arsenal = {p["codigo"] for p in productos}


# ════════════════════════════════════════════════════════════════════════════
# PASO 8: ARMAR LA SERIE DE CONSUMO MES A MES POR PRODUCTO Y POR CENTRO
# (solo productos del arsenal, para no inflar el archivo con ruido)
# ════════════════════════════════════════════════════════════════════════════
_log_paso(8, "Armando la serie de consumo mes a mes por producto y centro...")
centros = [nombre_centro(c) for c in cols_centro]

consumo_arsenal = ch[ch["Codigo"].isin(codigos_arsenal)]
agrupado = consumo_arsenal.groupby(["Codigo", "Mes"], as_index=False)[cols_centro].sum()

consumo_serie = []
for row in agrupado.to_dict("records"):
    valores = [round(float(row[c]), 1) for c in cols_centro]
    if any(valores):  # se omiten filas totalmente en cero para aligerar el archivo
        consumo_serie.append({"c": int(row["Codigo"]), "m": row["Mes"], "v": valores})

meses = sorted(agrupado["Mes"].unique().tolist())

print(f"Serie de consumo embebida: {len(consumo_serie)} filas (codigo x mes) "
      f"sobre {len(meses)} meses y {len(centros)} centros")


# ════════════════════════════════════════════════════════════════════════════
# PASO 8B: MÓDULO DE PRONÓSTICO DE DEMANDA (pronostico_demanda.py)
#
# Se pronostica cada combinación código+establecimiento y también el agregado
# de red por código, evaluando varios métodos con validación temporal y
# seleccionando el mejor por serie (ver pronostico_demanda.py para el detalle
# de calidad de datos, modelos y métricas). Es un cálculo pesado (miles de
# series), por lo que se cachea por huella de los archivos fuente: si
# Consumos_Historicos.xlsx y los Sal_Art_*.xlsx no cambiaron desde la última
# corrida, se reusa el resultado en vez de recalcular todo de nuevo.
#
# NOTA IMPORTANTE (transparencia de datos): los archivos de stock
# (Sal_Art_*.xlsx) no traen una columna que los vincule 1 a 1 con las
# columnas de establecimiento de Consumos_Historicos.xlsx, así que el stock
# NO se puede cruzar por establecimiento individual. Por eso el pronóstico de
# CONSUMO sí está disponible por establecimiento (para el gráfico histórico),
# pero los indicadores de ABASTECIMIENTO (stock de seguridad, punto de
# pedido, compra sugerida, cobertura, semáforo) solo pueden calcularse sobre
# los dos ámbitos de stock que sí existen en los datos: "Red completa"
# (Total Saldos) o "Solo BFC" (Saldo BFC, la bodega central que compra), y se
# recalculan en el navegador contra el pronóstico agregado de red.
# ════════════════════════════════════════════════════════════════════════════
_log_paso(9, "Calculando pronóstico de demanda y abastecimiento (puede tardar varios minutos si no hay caché)...")
RUTA_CACHE_PRONOSTICO = carpeta / ".pronostico_cache.pkl"


def _huella_archivos(rutas):
    partes = []
    for r in rutas:
        try:
            st = r.stat()
            partes.append(f"{r.name}:{st.st_mtime_ns}:{st.st_size}")
        except FileNotFoundError:
            partes.append(f"{r.name}:ausente")
    return hashlib.sha256("|".join(partes).encode()).hexdigest()


huella_actual = _huella_archivos([RUTA_CONSUMOS, *archivos_stock])
resultado_pronostico = None

if RUTA_CACHE_PRONOSTICO.exists():
    try:
        with open(RUTA_CACHE_PRONOSTICO, "rb") as f:
            cache_pronostico = pickle.load(f)
        if cache_pronostico.get("huella") == huella_actual:
            resultado_pronostico = cache_pronostico["resultado"]
            print("Pronósticos: usando caché (archivos fuente sin cambios desde la última corrida)")
    except Exception:
        resultado_pronostico = None

if resultado_pronostico is None:
    try:
        precios_por_codigo = dict(zip(df["Código"], df["Ult. Precio"]))
        print("Pronósticos: no hay caché válida, calculando (puede tardar varios minutos)...")
        resultado_pronostico = pronostico_demanda.generar_pronosticos(ch, cols_consumo, precios_por_codigo)
        with open(RUTA_CACHE_PRONOSTICO, "wb") as f:
            pickle.dump({"huella": huella_actual, "resultado": resultado_pronostico}, f)
    except Exception as exc:
        print(f"AVISO: el módulo de pronóstico falló ({exc}); el dashboard se genera sin esa sección.")
        resultado_pronostico = {"series": [], "calidad": {}, "clasificacion": {}}

indice_centro = {col: i for i, col in enumerate(cols_centro)}


def _compactar_pronosticos(series):
    filas = []
    for s in series:
        establecimiento = "RED" if s["establecimiento"] == "RED" else indice_centro.get(s["establecimiento"])
        if establecimiento is None:
            continue
        met = s["metricas"] or {}
        filas.append({
            "c": s["codigo"], "e": establecimiento, "mod": s["modelo"], "nc": s["nivel_confianza"],
            "wape": round(met["wape"], 4) if met.get("wape") is not None else None,
            "mase": round(met["mase"], 3) if met.get("mase") is not None else None,
            "sesgo": round(met["sesgo"], 2) if met.get("sesgo") is not None else None,
            "sig": round(float(s["sigma"]), 2),
            "f":  [round(float(v), 1) for v in s["forecast"]],
            "l8": [round(float(v), 1) for v in s["li80"]],
            "u8": [round(float(v), 1) for v in s["ls80"]],
            "l9": [round(float(v), 1) for v in s["li95"]],
            "u9": [round(float(v), 1) for v in s["ls95"]],
        })
    return filas


pronosticos_compactos = _compactar_pronosticos(resultado_pronostico["series"])
clasificacion_abc_xyz = {
    str(codigo): v for codigo, v in resultado_pronostico["clasificacion"].items()
}
calidad_datos = resultado_pronostico["calidad"]

METODOLOGIA = {
    "modelosDesc": pronostico_demanda.DESCRIPCION_MODELOS,
    "periodoEntrenamiento": f"{ch['Fecha'].min().strftime('%Y-%m')} a {fecha_max.strftime('%Y-%m')}",
    "horizontes": [3, 6, 12],
    "metricas": ["MAE", "RMSE", "WAPE", "MASE", "Sesgo del pronóstico"],
    "reglasHistoria": [
        f"Menos de {pronostico_demanda.MIN_MESES_NO_ESTACIONAL} meses de historia: "
        "\"historia insuficiente\", se usa un promedio simple claramente identificado.",
        f"Entre {pronostico_demanda.MIN_MESES_NO_ESTACIONAL} y "
        f"{pronostico_demanda.MIN_MESES_ESTACIONAL - 1} meses: solo modelos no estacionales.",
        f"Desde {pronostico_demanda.MIN_MESES_ESTACIONAL} meses: se habilitan también modelos "
        "estacionales (Holt-Winters, SARIMA, naive estacional).",
    ],
    "supuestos": [
        "Los meses sin registro dentro de la vida activa de un producto se tratan como 0 para "
        "el modelo, pero se cuentan y se muestran aparte (no se inventan valores).",
        "Los consumos negativos detectados se corrigen a 0 antes de modelar (error de origen).",
        "El lead time de reposición, el múltiplo de compra y el nivel de servicio son parámetros "
        "configurables en el dashboard (no existen en los datos fuente por producto/proveedor).",
        "La selección de modelo usa validación temporal walk-forward: nunca se entrena con datos "
        "posteriores al período que se está evaluando.",
        "SARIMA solo se evalúa sobre el agregado de red por producto (no por establecimiento), "
        "por ser el modelo más costoso de ajustar y el de menor aporte en series delgadas.",
    ],
    "limitaciones": [
        "No existe fecha de vencimiento/caducidad en los archivos fuente: el riesgo de "
        "vencimiento se muestra como \"No disponible\", no se calcula.",
        "No existe stock en tránsito real (compras ya despachadas por el proveedor) en los "
        "archivos fuente: se asume 0 salvo que el usuario lo indique manualmente en el dashboard.",
        "Los archivos de stock (Sal_Art_*) no tienen una columna que los vincule con las "
        "columnas de establecimiento de Consumos_Historicos: los indicadores de abastecimiento "
        "se calculan sobre el stock de toda la red o sobre BFC, no por punto de dispensación.",
        "La columna \"Categoría\" de los archivos de stock viene vacía en casi todos los "
        "productos; el filtro por categoría/programa quedará mayormente en \"Sin categoría\".",
        "Todas las cifras de pronóstico son estimaciones estadísticas sujetas a incertidumbre, "
        "no valores exactos.",
    ],
}

print(f"Calidad de datos: {calidad_datos.get('codigos_historia_insuficiente', 0)} códigos con "
      f"historia insuficiente, {calidad_datos.get('duplicados_codigo_fecha', 0)} filas "
      f"duplicadas, {calidad_datos.get('valores_negativos_corregidos', 0)} valores negativos "
      f"corregidos, {calidad_datos.get('outliers_detectados', 0)} outliers detectados")


# ════════════════════════════════════════════════════════════════════════════
# PASO 9: EMPAQUETAR TODO COMO JSON PARA EL FRONTEND
# ════════════════════════════════════════════════════════════════════════════
_log_paso(10, "Empaquetando los datos y generando el archivo HTML final del dashboard...")
DATA = {
    "meses": meses,
    "centros": centros,
    "productos": productos,
    "consumo": consumo_serie,
    "centrosStock": centros_stock,
    "saldoPorEstablecimiento": [
        {"centro": r["centro"], "unidades": int(r["unidades"]),
         "productos": int(r["productos"]), "valor": float(r["valor"])}
        for _, r in saldo_por_establecimiento.iterrows()
    ],
    "mesInicioDefecto": fecha_inicio.strftime("%Y-%m"),
    "mesFinDefecto": fecha_max.strftime("%Y-%m"),
    "generadoEl": datetime.now().strftime("%d-%m-%Y %H:%M"),
    "pronosticos": pronosticos_compactos,
    "clasificacion": clasificacion_abc_xyz,
    "calidadDatos": calidad_datos,
    "metodologia": METODOLOGIA,
}

# json.dumps con ensure_ascii=False para no reventar el archivo con \uXXXX;
# se escapan las barras de cierre de <script> para que ningún nombre de
# producto pueda cortar el bloque de script por accidente.
data_json = json.dumps(DATA, ensure_ascii=False, separators=(",", ":"))
data_json = data_json.replace("</", "<\\/")

plotly_js = get_plotlyjs()  # librería completa embebida: el HTML no depende de internet

print(f"Tamaño aprox. del JSON embebido: {len(data_json) / 1024 / 1024:.1f} MB")


# ════════════════════════════════════════════════════════════════════════════
# PASO 10: PLANTILLA HTML — estructura, estilos y lógica en el navegador
# ════════════════════════════════════════════════════════════════════════════
PLANTILLA = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Dashboard Farmacia — BFC</title>
<style>
:root{
    --urgente:#FF4444; --normal:#FFD966; --ok:#70AD47; --sinconsumo:#BFBFBF; --header:#1F4E79;
    --sobrestock:#4472C4; --banda80:rgba(31,78,121,0.22); --banda95:rgba(31,78,121,0.10);
}
*{ box-sizing:border-box; }
body{
    margin:0; padding:0 24px 40px; background:#F2F4F7; color:#1F2937;
    font-family:"Segoe UI",Arial,sans-serif;
}
header{ background:var(--header); color:#fff; margin:0 -24px 20px; padding:26px 24px; }
header h1{ margin:0; font-size:25px; }
header p{ margin:6px 0 0; font-size:13px; opacity:.85; }

.panel{ background:#fff; border-radius:10px; padding:16px; margin-bottom:22px; box-shadow:0 2px 6px rgba(0,0,0,.08); }
.panel h2{ margin:0 0 12px; font-size:16px; color:var(--header); }
.panel-doble{ display:grid; grid-template-columns:1fr 1fr; gap:22px; align-items:stretch; margin-bottom:22px; }
.panel-doble .panel{ margin-bottom:0; display:flex; flex-direction:column; }
.panel-doble .panel > div[id]{ flex:1; }
@media (max-width:980px){ .panel-doble{ grid-template-columns:1fr; } }

/* Paneles de gráficos: borde superior de color + encabezado opcional en línea con sus filtros */
.chart-panel{ border-top:4px solid var(--header); }
.chart-panel-head{ display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:10px; margin-bottom:12px; }
.chart-panel-head h2{ margin:0; }
.chart-panel .chart-filtros{ margin-bottom:0; }

/* Barra de filtros globales */
.filtros-bar{ display:flex; flex-wrap:wrap; align-items:flex-end; gap:18px; position:sticky; top:0; z-index:50; }
.filtro-grupo{ display:flex; flex-direction:column; gap:4px; font-size:12px; color:#555; }
.filtro-grupo label{ font-weight:600; }
.filtro-grupo select, .filtro-grupo input{
    padding:7px 9px; border:1px solid #D0D5DD; border-radius:6px; font-size:13px; min-width:150px;
}
.chk-group{ display:flex; gap:10px; flex-wrap:wrap; }
.chk-group label{ display:flex; align-items:center; gap:5px; font-weight:500; font-size:12px; color:#333; cursor:pointer; }
#btnReset{
    background:var(--header); color:#fff; border:none; border-radius:6px;
    padding:9px 16px; font-size:13px; cursor:pointer; height:34px;
}
#btnReset:hover{ opacity:.9; }

/* KPIs */
.kpis{ display:grid; grid-template-columns:repeat(auto-fit,minmax(170px,1fr)); gap:16px; margin-bottom:22px; }
.kpi-card{ background:#fff; border-radius:10px; border-top:6px solid #ccc; padding:18px; text-align:center; box-shadow:0 2px 6px rgba(0,0,0,.08); }
.kpi-valor{ font-size:32px; font-weight:700; line-height:1; }
.kpi-etiqueta{ margin-top:8px; font-size:12px; font-weight:600; letter-spacing:.03em; color:#555; }

/* Badges de alerta */
.badge{ display:inline-block; padding:4px 12px; border-radius:20px; color:#fff; font-weight:700; font-size:12px; }
.badge-URGENTE{ background:var(--urgente); }
.badge-NORMAL{ background:var(--normal); color:#5c4a00; }
.badge-OK{ background:var(--ok); }
.badge-SINCONSUMO,.badge-SIN.CONSUMO{ background:var(--sinconsumo); color:#333; }
.badge-ROJO{ background:var(--urgente); }
.badge-AMARILLO{ background:var(--normal); color:#5c4a00; }
.badge-VERDE{ background:var(--ok); }
.badge-AZUL{ background:var(--sobrestock); }

/* Módulo de pronóstico */
.aviso-incertidumbre{ font-size:12px; color:#7a5c00; background:#FFF7DE; border:1px solid #FFE79A; border-radius:6px; padding:8px 12px; margin:0 0 14px; }
.leyenda-semaforo{ display:flex; gap:16px; flex-wrap:wrap; font-size:12px; color:#444; margin-bottom:12px; }
.leyenda-semaforo span{ display:flex; align-items:center; gap:6px; }
.punto{ display:inline-block; width:11px; height:11px; border-radius:50%; }
.metodologia-lista{ margin:6px 0 16px; padding-left:20px; font-size:13px; color:#333; line-height:1.6; }
.metodologia-grid{ display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:12px; margin-bottom:14px; }
.metodologia-dato{ background:#F7F9FC; border-radius:8px; padding:10px 12px; }
.metodologia-dato b{ display:block; color:var(--header); font-size:13px; }
.metodologia-dato span{ font-size:12px; color:#666; }
#btnDescargarCSV, #btnDescargarExcel, #btnDescargarSolicitud{
    background:#fff; color:var(--header); border:1px solid var(--header); border-radius:6px;
    padding:8px 14px; font-size:13px; cursor:pointer; margin-right:10px; margin-top:8px;
}
#btnDescargarCSV:hover, #btnDescargarExcel:hover, #btnDescargarSolicitud:hover{ background:#EEF3FA; }
#solHorizonte{ padding:6px 9px; border:1px solid #D0D5DD; border-radius:6px; font-size:13px; margin-left:6px; }

/* Sub-filtros de gráfico de línea */
.chart-filtros{ display:flex; gap:14px; flex-wrap:wrap; margin-bottom:10px; }
.chart-filtros select, .chart-filtros input{ padding:6px 9px; border:1px solid #D0D5DD; border-radius:6px; font-size:13px; min-width:220px; }

/* Buscador / ficha */
#buscadorInput{ width:100%; max-width:420px; padding:10px 12px; border:1px solid #D0D5DD; border-radius:8px; font-size:14px; }
.ficha-vacia{ color:#888; font-style:italic; padding:20px 0; }
.ficha-header{ display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:10px; margin:16px 0; }
.ficha-titulo{ font-size:18px; font-weight:700; color:var(--header); }
.ficha-sub{ color:#666; font-size:13px; }
.ficha-kpis{ display:grid; grid-template-columns:repeat(auto-fit,minmax(140px,1fr)); gap:10px; margin-bottom:16px; }
.ficha-kpi{ background:#F7F9FC; border-radius:8px; padding:10px; text-align:center; }
.ficha-kpi b{ display:block; font-size:18px; color:var(--header); }
.ficha-kpi span{ font-size:11px; color:#666; }
.ficha-graficos{ display:grid; grid-template-columns:1fr 1fr; gap:18px; }
@media (max-width:980px){ .ficha-graficos{ grid-template-columns:1fr; } }

/* Tabla */
#tablaSearch{ width:100%; max-width:360px; padding:8px 10px; border:1px solid #D0D5DD; border-radius:6px; font-size:13px; margin-bottom:12px; }
.tabla-wrap{ max-height:600px; overflow:auto; border:1px solid #E5E7EB; border-radius:8px; }
table{ border-collapse:collapse; width:100%; font-size:12.5px; }
thead th{
    position:sticky; top:0; background:var(--header); color:#fff; padding:10px 8px;
    text-align:left; cursor:pointer; user-select:none; white-space:nowrap;
}
thead th:hover{ background:#173a5c; }
thead th .arrow{ opacity:.6; margin-left:3px; }
tbody td{ padding:8px; border-bottom:1px solid #EEF0F3; white-space:nowrap; }
tbody tr:hover{ background:#F5F8FC; }
.num{ text-align:right; font-variant-numeric:tabular-nums; }
.contador-tabla{ font-size:12px; color:#666; margin-bottom:8px; }

/* Barra de magnitud (un solo tono, ancho proporcional) para comparar totales
   dentro de una misma columna sin necesitar un gráfico aparte */
.celda-barra{ display:flex; align-items:center; gap:8px; justify-content:flex-end; }
.barra-mini{ position:relative; width:70px; height:8px; background:#EEF3FA; border-radius:4px; overflow:hidden; flex-shrink:0; }
.barra-mini > div{ position:absolute; top:0; left:0; height:100%; background:var(--header); border-radius:4px; }
.tendencia-sube{ color:#B03A2E; font-weight:600; }
.tendencia-baja{ color:#1E7145; font-weight:600; }
.tendencia-estable{ color:#666; font-weight:600; }

footer{ text-align:center; color:#888; font-size:12px; margin-top:10px; }
</style>
<script type="text/javascript">__PLOTLYJS__</script>
</head>
<body>

<header>
    <h1>Dashboard Farmacia — Droguería BFC</h1>
    <p id="headerMeta"></p>
</header>

<section class="panel filtros-bar">
    <div class="filtro-grupo" style="flex:1 1 280px;">
        <label>Medicamento (código o nombre)</label>
        <input id="buscadorInput" list="datalistMedicamentos" placeholder="Todos los medicamentos...">
        <datalist id="datalistMedicamentos"></datalist>
    </div>
    <div class="filtro-grupo">
        <label>Mes desde</label>
        <select id="fMesDesde"></select>
    </div>
    <div class="filtro-grupo">
        <label>Mes hasta</label>
        <select id="fMesHasta"></select>
    </div>
    <div class="filtro-grupo">
        <label>Centro</label>
        <select id="fCentro"><option value="__ALL__">Todos los centros</option></select>
    </div>
    <div class="filtro-grupo">
        <label>Alerta de stock</label>
        <div class="chk-group" id="fAlerta"></div>
    </div>
    <button id="btnReset">Limpiar filtros</button>
</section>

<section class="kpis" id="kpiSection"></section>

<div class="panel">
    <h2>Saldo por establecimiento</h2>
    <p class="ficha-sub" id="saldoEstablecimientoSub">Stock actual de toda la red, sumando todos los productos del arsenal, por bodega/centro.</p>
    <div class="tabla-wrap" style="max-height:420px;">
        <table id="tablaSaldoEstablecimiento">
            <thead><tr id="saldoEstablecimientoHead">
                <th>Establecimiento</th><th class="num">Unidades en stock</th>
                <th class="num">Productos distintos</th><th class="num">Valorización</th>
            </tr></thead>
            <tbody id="saldoEstablecimientoBody"></tbody>
        </table>
    </div>
</div>

<div class="panel chart-panel">
    <div class="chart-panel-head">
        <h2>Consumo histórico mes a mes</h2>
        <div class="chart-filtros">
            <select id="lineCentro"><option value="__ALL__">Todos los centros</option></select>
        </div>
    </div>
    <div id="chartLinea"></div>
</div>

<div class="panel chart-panel">
    <h2>Ranking Top 20 — medicamentos más consumidos</h2>
    <div id="chartRanking"></div>
</div>

<div class="panel-doble">
    <div class="panel chart-panel"><h2>Comparación de consumo entre centros (apilado)</h2><div id="chartStack"></div></div>
    <div class="panel chart-panel"><h2>Proyección de consumo a 3 y 6 meses</h2><div id="chartProy"></div></div>
</div>

<div class="panel">
    <h2>Consumo por establecimiento — últimos 6 meses</h2>
    <p class="ficha-sub" id="consumoEstablecimientoSub"></p>
    <div class="tabla-wrap" style="max-height:460px;">
        <table id="tablaConsumoEstablecimiento">
            <thead><tr>
                <th>Establecimiento</th>
                <th class="num">Total 6M</th>
                <th class="num">Promedio mensual</th>
                <th class="num">Máximo mensual</th>
                <th class="num">Mínimo mensual</th>
                <th class="num">Participación</th>
                <th class="num">Tendencia</th>
            </tr></thead>
            <tbody id="consumoEstablecimientoBody"></tbody>
        </table>
    </div>
</div>

<div class="panel chart-panel">
    <h2>Comparación mensual de consumo entre establecimientos</h2>
    <p class="ficha-sub" id="comparacionMensualSub"></p>
    <div id="chartComparacionMensual"></div>
</div>

<div class="panel">
    <h2>Ficha del medicamento seleccionado</h2>
    <div id="fichaContenido" class="ficha-vacia">Escribe un código o nombre en el buscador de arriba para ver su ficha completa.</div>
</div>

<div class="panel">
    <h2>Tabla resumen</h2>
    <input id="tablaSearch" placeholder="Buscar por código, medicamento o proveedor...">
    <div class="contador-tabla" id="tablaContador"></div>
    <div class="tabla-wrap">
        <table id="tablaResumen">
            <thead><tr id="tablaHead"></tr></thead>
            <tbody id="tablaBody"></tbody>
        </table>
    </div>
</div>

<div class="panel" style="border-left:6px solid var(--header)">
    <h2>Pronóstico de demanda y abastecimiento</h2>
    <p class="aviso-incertidumbre">
        Estas cifras son estimaciones estadísticas obtenidas con modelos de series de tiempo
        y están sujetas a incertidumbre — no son valores exactos. Revisa la sección
        "Metodología" al final de la página para conocer los modelos, supuestos y limitaciones.
    </p>
</div>

<section class="panel filtros-bar" id="filtrosPronostico">
    <div class="filtro-grupo"><label>Horizonte</label>
        <select id="pHorizonte">
            <option value="3">3 meses</option>
            <option value="6">6 meses</option>
            <option value="12" selected>12 meses</option>
        </select>
    </div>
    <div class="filtro-grupo"><label>Nivel de servicio</label>
        <select id="pNivelServicio">
            <option value="90">90%</option>
            <option value="95" selected>95%</option>
            <option value="97.5">97.5%</option>
            <option value="99">99%</option>
        </select>
    </div>
    <div class="filtro-grupo"><label>Ámbito de stock</label>
        <select id="pAmbitoStock">
            <option value="red" selected>Red completa</option>
            <option value="bfc">Solo BFC</option>
        </select>
    </div>
    <div class="filtro-grupo"><label>Lead time (días)</label>
        <input type="number" id="pLeadTime" value="15" min="0" step="1">
    </div>
    <div class="filtro-grupo"><label>Múltiplo de compra</label>
        <input type="number" id="pMultiplo" value="1" min="1" step="1">
    </div>
    <div class="filtro-grupo"><label>Stock en tránsito (manual)</label>
        <input type="number" id="pStockTransito" value="0" min="0" step="1">
    </div>
    <div class="filtro-grupo"><label>Método de pronóstico</label>
        <select id="pMetodo"><option value="__ALL__">Todos</option></select>
    </div>
    <div class="filtro-grupo"><label>Clasificación ABC</label>
        <select id="pABC">
            <option value="__ALL__">Todas</option><option value="A">A</option>
            <option value="B">B</option><option value="C">C</option>
        </select>
    </div>
    <div class="filtro-grupo"><label>Clasificación XYZ</label>
        <select id="pXYZ">
            <option value="__ALL__">Todas</option><option value="X">X</option>
            <option value="Y">Y</option><option value="Z">Z</option><option value="N/D">N/D</option>
        </select>
    </div>
    <div class="filtro-grupo"><label>Categoría / Programa</label>
        <select id="pCategoria"><option value="__ALL__">Todas</option></select>
    </div>
    <button id="btnResetPronostico">Limpiar filtros de pronóstico</button>
</section>

<section class="kpis" id="kpiPronostico"></section>

<div class="panel chart-panel">
    <div class="chart-panel-head">
        <h2>Consumo histórico, pronóstico e intervalos de predicción</h2>
        <div class="chart-filtros">
            <span class="ficha-sub">Usa el buscador de medicamento de arriba (barra de filtros) para elegir el producto.</span>
        </div>
    </div>
    <div id="chartPronostico"></div>
    <div id="infoModeloSerie" class="ficha-sub" style="margin-top:8px"></div>
</div>

<div class="panel">
    <h2>Tabla de priorización por riesgo de abastecimiento</h2>
    <div class="leyenda-semaforo">
        <span><i class="punto" style="background:var(--urgente)"></i> Rojo: quiebre probable</span>
        <span><i class="punto" style="background:var(--normal)"></i> Amarillo: cobertura insuficiente</span>
        <span><i class="punto" style="background:var(--ok)"></i> Verde: cobertura adecuada</span>
        <span><i class="punto" style="background:var(--sobrestock)"></i> Azul: posible sobrestock</span>
    </div>
    <input id="pTablaSearch" placeholder="Buscar por código, medicamento o proveedor...">
    <div class="contador-tabla" id="pTablaContador"></div>
    <div class="tabla-wrap">
        <table id="tablaPronostico">
            <thead><tr id="pTablaHead"></tr></thead>
            <tbody id="pTablaBody"></tbody>
        </table>
    </div>
    <div>
        <button id="btnDescargarCSV">Descargar CSV</button>
        <button id="btnDescargarExcel">Descargar Excel</button>
    </div>
</div>

<div class="panel" id="panelSolicitud">
    <h2>Solicitud de compra</h2>
    <p class="ficha-sub">
        Genera la solicitud de compra (mismo formato que el archivo modelo SOLICITUD) con los
        productos en riesgo <b>ROJO</b> (quiebre probable) o <b>AMARILLO</b> (cobertura insuficiente)
        para el horizonte elegido, usando los parámetros de abastecimiento configurados arriba
        (nivel de servicio, lead time, múltiplo, stock en tránsito, ámbito de stock).
    </p>
    <div class="chart-filtros" style="align-items:center;">
        <label style="font-size:13px;color:#444;">Horizonte de la solicitud
            <select id="solHorizonte">
                <option value="3">3 meses</option>
                <option value="6" selected>6 meses</option>
                <option value="12">12 meses</option>
            </select>
        </label>
        <button id="btnDescargarSolicitud">Descargar Solicitud (Excel)</button>
    </div>
    <div class="contador-tabla" id="solContador"></div>
</div>

<div class="panel" id="panelGeneradorPedido">
    <h2>Generador de Pedido — Horizonte Variable</h2>
    <p class="ficha-sub">
        Lista los medicamentos con <b>quiebre de stock proyectado dentro del horizonte</b> elegido
        (según el pronóstico de demanda y los parámetros de abastecimiento configurados arriba:
        nivel de servicio, lead time, múltiplo, stock en tránsito, ámbito de stock) y calcula la
        cantidad a pedir para cubrir esa ventana de tiempo. A diferencia de la solicitud por
        semáforo, aquí el horizonte se ajusta libremente en días, no en tramos fijos de 3/6/12 meses.
    </p>
    <div class="chart-filtros" style="align-items:center; gap:16px;">
        <label style="font-size:13px;color:#444;">
            Horizonte del pedido: <b id="pedHorizonteLabel">90 días (~3.0 meses)</b><br>
            <input type="range" id="pedHorizonteDias" min="7" max="365" step="1" value="90"
                   style="width:280px; vertical-align:middle; margin-top:4px;">
        </label>
    </div>
    <section class="kpis" id="kpiGeneradorPedido"></section>
    <input id="pedTablaSearch" placeholder="Buscar por código, medicamento o proveedor...">
    <div class="contador-tabla" id="pedTablaContador"></div>
    <div class="tabla-wrap">
        <table id="tablaGeneradorPedido">
            <thead><tr id="pedTablaHead"></tr></thead>
            <tbody id="pedTablaBody"></tbody>
        </table>
    </div>
    <div>
        <button id="btnDescargarPedido">Descargar Pedido (Excel)</button>
    </div>
</div>

<div class="panel" id="panelMetodologia">
    <h2>Metodología</h2>
    <div class="metodologia-grid" id="metodologiaDatos"></div>
    <p><b>Reglas de historia mínima:</b></p>
    <ul class="metodologia-lista" id="metodologiaReglas"></ul>
    <p><b>Métricas de desempeño usadas (nunca MAPE):</b> <span id="metodologiaMetricas"></span></p>
    <p><b>Supuestos:</b></p>
    <ul class="metodologia-lista" id="metodologiaSupuestos"></ul>
    <p><b>Limitaciones:</b></p>
    <ul class="metodologia-lista" id="metodologiaLimitaciones"></ul>
    <p><b>Calidad de datos detectada en la última corrida:</b></p>
    <ul class="metodologia-lista" id="metodologiaCalidad"></ul>
</div>

<footer>Generado automáticamente por generar_dashboard_completo.py — no requiere conexión a internet</footer>

<script type="text/javascript">
const DATA = __DATA_JSON__;

/* ── Índices auxiliares ──────────────────────────────────────────────── */
const centros = DATA.centros;
const centroIndex = {}; centros.forEach((c,i)=>centroIndex[c]=i);
const meses = DATA.meses;
const productos = DATA.productos;
const productoByCodigo = {}; productos.forEach(p=>productoByCodigo[p.codigo]=p);
const consumo = DATA.consumo;
const saldoPorEstablecimiento = DATA.saldoPorEstablecimiento;
const centrosStock = DATA.centrosStock;

const consumoByCodigo = {};
consumo.forEach(row=>{
    (consumoByCodigo[row.c] ||= []).push(row);
});
Object.values(consumoByCodigo).forEach(arr=>arr.sort((a,b)=>a.m.localeCompare(b.m)));

const ORDEN_ALERTA = ["URGENTE","NORMAL","OK","SIN CONSUMO"];
const COLOR_ALERTA = {URGENTE:"#FF4444", NORMAL:"#FFD966", OK:"#70AD47", "SIN CONSUMO":"#BFBFBF"};

function normaliza(s){
    return String(s ?? "").toLowerCase().normalize("NFD").replace(/[̀-ͯ]/g,"");
}
function fmtNum(v){ return new Intl.NumberFormat("es-CL",{maximumFractionDigits:0}).format(v || 0); }
function fmtCLP(v){ return "$" + fmtNum(v); }
function fmtCompacto(v){
    v = v || 0;
    if (Math.abs(v) >= 1e9) return "$" + (v/1e9).toFixed(2).replace(".",",") + " mil M";
    if (Math.abs(v) >= 1e6) return "$" + (v/1e6).toFixed(1).replace(".",",") + " M";
    if (Math.abs(v) >= 1e3) return "$" + (v/1e3).toFixed(0) + " mil";
    return fmtCLP(v);
}

/* ── Poblar controles de filtro ──────────────────────────────────────── */
const fMesDesde = document.getElementById("fMesDesde");
const fMesHasta = document.getElementById("fMesHasta");
const fCentro   = document.getElementById("fCentro");
const fAlertaBox = document.getElementById("fAlerta");
const lineCentro   = document.getElementById("lineCentro");
const datalistMed  = document.getElementById("datalistMedicamentos");
const buscadorInput = document.getElementById("buscadorInput");

/* Medicamento seleccionado globalmente (código o null = ninguno / todos).
   Un único buscador maneja la ficha, el gráfico de línea, el desglose por
   centro, el resaltado en ranking/proyección y el filtro de la tabla. */
let medicamentoSel = null;

meses.forEach(m=>{
    fMesDesde.appendChild(new Option(m, m));
    fMesHasta.appendChild(new Option(m, m));
});
fMesDesde.value = DATA.mesInicioDefecto;
fMesHasta.value = DATA.mesFinDefecto;

centros.forEach(c=>{
    fCentro.appendChild(new Option(c, c));
    lineCentro.appendChild(new Option(c, c));
});

ORDEN_ALERTA.forEach(niv=>{
    const id = "chk_" + niv.replace(/\s+/g,"");
    const lbl = document.createElement("label");
    lbl.innerHTML = `<input type="checkbox" id="${id}" data-alerta="${niv}" checked> ${niv}`;
    fAlertaBox.appendChild(lbl);
});

const medicamentosOrdenados = productos.slice().sort((a,b)=>a.medicamento.localeCompare(b.medicamento));
medicamentosOrdenados.forEach(p=>{
    datalistMed.appendChild(new Option(`${p.codigo} - ${p.medicamento}`));
});

document.getElementById("headerMeta").textContent =
    `${productos.length} productos del arsenal | Historial: ${meses[0]} a ${meses[meses.length-1]} | Generado el ${DATA.generadoEl}`;

/* ── Helpers de filtro global ────────────────────────────────────────── */
function getAlertaSet(){
    const set = new Set();
    fAlertaBox.querySelectorAll("input[type=checkbox]").forEach(chk=>{ if(chk.checked) set.add(chk.dataset.alerta); });
    return set;
}
function productosFiltrados(){
    const set = getAlertaSet();
    return productos.filter(p=>set.has(p.alerta));
}
function rangoMeses(){
    let desde = fMesDesde.value, hasta = fMesHasta.value;
    if (desde > hasta) [desde, hasta] = [hasta, desde];
    return [desde, hasta];
}

/* ── Estilo común para todos los gráficos Plotly ─────────────────────── */
const PALETA_CENTROS = ["#1F4E79","#70AD47","#FFB347","#C0504D","#4472C4","#8064A2",
                         "#2E9E8F","#D98880","#59788E","#B08968","#6A8EAE","#9BBB59"];
const PLOTLY_CONFIG = { displaylogo:false, responsive:true,
    modeBarButtonsToRemove:["lasso2d","select2d","autoScale2d"] };

function layoutBase(extra){
    // Object.assign de dos pasadas: la 2da pasada vuelve a fusionar
    // xaxis/yaxis/margin para que un "extra.xaxis" parcial (p.ej. solo title)
    // no borre el gridcolor/automargin por defecto de la 1ra pasada.
    return Object.assign({
        font: { family:"Segoe UI, Arial, sans-serif", size:12, color:"#333" },
        paper_bgcolor: "#fff", plot_bgcolor: "#fff",
        hoverlabel: { bgcolor:"#fff", bordercolor:"#D0D5DD", font:{color:"#1F2937", size:12} },
        margin: { l:50, r:20, t:20, b:40 },
        xaxis: { gridcolor:"#EEF0F3", zerolinecolor:"#E5E7EB", automargin:true },
        yaxis: { gridcolor:"#EEF0F3", zerolinecolor:"#E5E7EB", automargin:true },
    }, extra, {
        xaxis: Object.assign({ gridcolor:"#EEF0F3", zerolinecolor:"#E5E7EB", automargin:true }, extra.xaxis || {}),
        yaxis: Object.assign({ gridcolor:"#EEF0F3", zerolinecolor:"#E5E7EB", automargin:true }, extra.yaxis || {}),
        margin: Object.assign({ l:50, r:20, t:20, b:40 }, extra.margin || {}),
    });
}

/* ── KPIs ─────────────────────────────────────────────────────────────── */
function renderKPIs(){
    const activos = productosFiltrados();
    const cont = document.getElementById("kpiSection");
    const valorTotal = activos.reduce((s,p)=>s+p.valorStockTotal, 0);
    const conteos = {}; ORDEN_ALERTA.forEach(n=>conteos[n]=0);
    activos.forEach(p=>{ conteos[p.alerta] = (conteos[p.alerta]||0)+1; });

    let html = `
      <div class="kpi-card" style="border-top-color:var(--header)">
        <div class="kpi-valor" style="color:var(--header)">${fmtNum(activos.length)}</div>
        <div class="kpi-etiqueta">TOTAL PRODUCTOS</div>
      </div>
      <div class="kpi-card" style="border-top-color:var(--header)" title="${fmtCLP(valorTotal)}">
        <div class="kpi-valor" style="color:var(--header)">${fmtCompacto(valorTotal)}</div>
        <div class="kpi-etiqueta">VALOR STOCK RED</div>
      </div>`;
    ORDEN_ALERTA.forEach(niv=>{
        html += `
      <div class="kpi-card" style="border-top-color:${COLOR_ALERTA[niv]}">
        <div class="kpi-valor" style="color:${COLOR_ALERTA[niv]}">${fmtNum(conteos[niv])}</div>
        <div class="kpi-etiqueta">${niv}</div>
      </div>`;
    });
    cont.innerHTML = html;
}

/* ── Saldo por establecimiento: total de la red, o desglose del medicamento
   seleccionado (independiente de los filtros de mes/centro/alerta) ─────── */
function renderSaldoEstablecimiento(){
    const tbody = document.getElementById("saldoEstablecimientoBody");
    const sub   = document.getElementById("saldoEstablecimientoSub");
    const head  = document.getElementById("saldoEstablecimientoHead");
    const p = medicamentoSel !== null ? productoByCodigo[medicamentoSel] : null;

    let filas, colProductos = true;
    if (p){
        sub.textContent = `Saldo de "${p.medicamento}" (código ${p.codigo}) por bodega/centro.`;
        head.innerHTML = `<th>Establecimiento</th><th class="num">Unidades en stock</th><th class="num">Valorización</th>`;
        colProductos = false;
        filas = centrosStock
            .map((c,i)=>({ centro:c, unidades:p.saldoCentros[i], valor:p.saldoCentros[i]*p.ultPrecio }))
            .filter(r=>r.unidades > 0)
            .sort((a,b)=>b.unidades - a.unidades);
    } else {
        sub.textContent = "Stock actual de toda la red, sumando todos los productos del arsenal, por bodega/centro.";
        head.innerHTML = `<th>Establecimiento</th><th class="num">Unidades en stock</th><th class="num">Productos distintos</th><th class="num">Valorización</th>`;
        filas = saldoPorEstablecimiento;
    }

    const totalUnidades = filas.reduce((s,r)=>s+r.unidades, 0);
    const totalValor     = filas.reduce((s,r)=>s+r.valor, 0);

    if (!filas.length){
        tbody.innerHTML = `<tr><td colspan="${colProductos?4:3}">Sin stock registrado en ningún establecimiento.</td></tr>`;
        return;
    }

    let html = filas.map(r => `
        <tr>
            <td>${r.centro}</td>
            <td class="num">${fmtNum(r.unidades)}</td>
            ${colProductos ? `<td class="num">${fmtNum(r.productos)}</td>` : ""}
            <td class="num">${fmtCLP(r.valor)}</td>
        </tr>`).join("");
    html += `
        <tr style="font-weight:bold; background:#F5F8FC;">
            <td>TOTAL RED</td>
            <td class="num">${fmtNum(totalUnidades)}</td>
            ${colProductos ? `<td class="num"></td>` : ""}
            <td class="num">${fmtCLP(totalValor)}</td>
        </tr>`;
    tbody.innerHTML = html;
}

/* ── Gráfico de línea: consumo mes a mes filtrable por centro/artículo ─── */
function renderLineChart(){
    const [desde, hasta] = rangoMeses();
    const mesesRango = meses.filter(m=>m>=desde && m<=hasta);
    const centroSel = lineCentro.value;

    let codigosConsiderar;
    if (medicamentoSel === null) {
        codigosConsiderar = new Set(productosFiltrados().map(p=>p.codigo));
    } else {
        codigosConsiderar = new Set([medicamentoSel]);
    }

    const totales = {}; mesesRango.forEach(m=>totales[m]=0);
    consumo.forEach(row=>{
        if (row.m < desde || row.m > hasta) return;
        if (!codigosConsiderar.has(row.c)) return;
        const val = centroSel === "__ALL__" ? row.v.reduce((a,b)=>a+b,0) : (row.v[centroIndex[centroSel]] || 0);
        totales[row.m] += val;
    });

    const titulo = medicamentoSel === null
        ? `Consumo total ${centroSel === "__ALL__" ? "de la red" : "en " + centroSel} (productos con la alerta seleccionada)`
        : `Consumo de "${productoByCodigo[medicamentoSel]?.medicamento ?? medicamentoSel}" ${centroSel === "__ALL__" ? "(toda la red)" : "en " + centroSel}`;

    Plotly.react("chartLinea", [{
        x: mesesRango, y: mesesRango.map(m=>totales[m]),
        mode: "lines+markers", type: "scatter", line:{color:"#1F4E79", width:3}, marker:{size:6, color:"#1F4E79"},
        fill: "tozeroy", fillcolor: "rgba(31,78,121,0.10)", hovertemplate: "%{x}<br><b>%{y:,.0f} unid.</b><extra></extra>",
    }], layoutBase({
        title: { text:titulo, font:{size:14} }, height:380, hovermode:"x unified",
        margin:{t:44}, xaxis:{title:"Mes"}, yaxis:{title:"Unidades consumidas", rangemode:"tozero"},
    }), PLOTLY_CONFIG);
}

/* ── Ranking Top 20 medicamentos más consumidos ──────────────────────── */
function renderRanking(){
    const [desde, hasta] = rangoMeses();
    const alertaSet = getAlertaSet();
    const centroSel = fCentro.value;
    const sumas = {};

    consumo.forEach(row=>{
        if (row.m < desde || row.m > hasta) return;
        const p = productoByCodigo[row.c];
        if (!p || !alertaSet.has(p.alerta)) return;
        const val = centroSel === "__ALL__" ? row.v.reduce((a,b)=>a+b,0) : (row.v[centroIndex[centroSel]] || 0);
        sumas[row.c] = (sumas[row.c] || 0) + val;
    });

    let arr = Object.entries(sumas)
        .map(([c,v])=>({codigo:+c, nombre: productoByCodigo[c]?.medicamento || ("Código " + c), total: v}))
        .filter(r=>r.total > 0)
        .sort((a,b)=>b.total - a.total)
        .slice(0, 20)
        .reverse();

    const bordeSel = arr.map(r=> r.codigo === medicamentoSel ? "#1F2937" : "rgba(0,0,0,0)");
    const anchoSel  = arr.map(r=> r.codigo === medicamentoSel ? 3 : 0);

    Plotly.react("chartRanking", [{
        x: arr.map(r=>r.total), y: arr.map(r=>r.nombre.slice(0,42)),
        type:"bar", orientation:"h", marker:{color:"#70AD47", line:{color:bordeSel, width:anchoSel}},
        text: arr.map(r=>fmtNum(r.total)), textposition:"outside", cliponaxis:false,
        hovertemplate: "%{y}<br><b>%{x:,.0f} unid.</b><extra></extra>",
    }], layoutBase({
        margin:{l:10,r:50,t:10,b:40}, height: Math.max(460, 26*arr.length),
        xaxis:{title:"Unidades consumidas"}, yaxis:{automargin:true},
    }), PLOTLY_CONFIG);
}

/* ── Comparación entre centros (apilado) ─────────────────────────────── */
function renderStack(){
    const [desde, hasta] = rangoMeses();
    const alertaSet = getAlertaSet();
    const mesesRango = meses.filter(m=>m>=desde && m<=hasta);
    const totalesPorCentro = new Array(centros.length).fill(0);
    const matriz = {}; mesesRango.forEach(m=>matriz[m] = new Array(centros.length).fill(0));

    consumo.forEach(row=>{
        if (row.m < desde || row.m > hasta) return;
        if (medicamentoSel !== null) {
            if (row.c !== medicamentoSel) return;
        } else {
            const p = productoByCodigo[row.c];
            if (!p || !alertaSet.has(p.alerta)) return;
        }
        row.v.forEach((val,i)=>{ matriz[row.m][i] += val; totalesPorCentro[i] += val; });
    });

    const topIdx = totalesPorCentro
        .map((v,i)=>[i,v]).sort((a,b)=>b[1]-a[1]).slice(0,8).map(x=>x[0]);

    const trazas = topIdx.map((i,pos)=>({
        x: mesesRango, y: mesesRango.map(m=>matriz[m][i]),
        name: centros[i], type:"bar", marker:{color: PALETA_CENTROS[pos % PALETA_CENTROS.length]},
        hovertemplate: "%{fullData.name}<br><b>%{y:,.0f} unid.</b><extra></extra>",
    }));

    Plotly.react("chartStack", trazas, layoutBase({
        title: medicamentoSel !== null
            ? { text:`Desglose por centro — ${productoByCodigo[medicamentoSel]?.medicamento ?? medicamentoSel}`, font:{size:13} }
            : undefined,
        barmode:"stack", height:440, hovermode:"x unified",
        margin:{b:70, t: medicamentoSel !== null ? 40 : 20},
        xaxis:{title:"Mes"}, yaxis:{title:"Unidades consumidas"},
        legend:{orientation:"h", y:-0.28, font:{size:10}},
    }), PLOTLY_CONFIG);
}

/* ── Consumo por establecimiento, ventana fija de los últimos 6 meses con
   datos (no se mueve con los filtros de mes, para que siempre responda la
   misma pregunta: "¿cómo viene consumiendo cada establecimiento últimamente?").
   Sin medicamento seleccionado, agrega los productos con la alerta marcada;
   con uno seleccionado, muestra solo su consumo por centro. ─────────────── */
function renderConsumoEstablecimientos(){
    const ultimos6 = meses.slice(-6);
    const sub  = document.getElementById("consumoEstablecimientoSub");
    const tbody = document.getElementById("consumoEstablecimientoBody");
    const rango = ultimos6.length ? `${ultimos6[0]} a ${ultimos6[ultimos6.length - 1]}` : "sin datos";

    let codigosConsiderar;
    if (medicamentoSel !== null) {
        const p = productoByCodigo[medicamentoSel];
        codigosConsiderar = new Set([medicamentoSel]);
        sub.textContent = `Consumo de "${p?.medicamento ?? medicamentoSel}" por establecimiento (${rango}).`;
    } else {
        codigosConsiderar = new Set(productosFiltrados().map(p=>p.codigo));
        sub.textContent = `Consumo de los productos del arsenal con la alerta seleccionada, por establecimiento (${rango}).`;
    }

    const porMesCentro = ultimos6.map(()=> new Array(centros.length).fill(0));
    consumo.forEach(row=>{
        const idxMes = ultimos6.indexOf(row.m);
        if (idxMes === -1 || !codigosConsiderar.has(row.c)) return;
        row.v.forEach((val,i)=>{ porMesCentro[idxMes][i] += val; });
    });

    const nMeses = ultimos6.length || 1;
    const mitad = Math.max(1, Math.floor(ultimos6.length / 2));
    let stats = centros.map((centro,i)=>{
        const valores = porMesCentro.map(fila=>fila[i]);
        const total = valores.reduce((a,b)=>a+b, 0);
        const primeraMitad = valores.slice(0, mitad).reduce((a,b)=>a+b, 0);
        const segundaMitad = valores.slice(mitad).reduce((a,b)=>a+b, 0);
        const tendencia = primeraMitad > 0 ? ((segundaMitad - primeraMitad) / primeraMitad) * 100 : null;
        return {
            centro, total, promedio: total / nMeses,
            max: valores.length ? Math.max(...valores) : 0,
            min: valores.length ? Math.min(...valores) : 0,
            tendencia,
        };
    }).filter(s=>s.total > 0).sort((a,b)=>b.total - a.total);

    if (!stats.length){
        tbody.innerHTML = `<tr><td colspan="7">Sin consumo registrado en los últimos 6 meses para este filtro.</td></tr>`;
        return;
    }

    const totalGeneral = stats.reduce((s,r)=>s+r.total, 0);
    const maxTotal = stats[0].total;

    tbody.innerHTML = stats.map(s=>{
        const participacion = totalGeneral ? (s.total / totalGeneral * 100) : 0;
        const anchoBarra = maxTotal ? (s.total / maxTotal * 100) : 0;
        let tendCelda = '<span style="color:#888">Sin mes previo</span>';
        if (s.tendencia !== null){
            const clase = Math.abs(s.tendencia) < 5 ? "tendencia-estable" : (s.tendencia > 0 ? "tendencia-sube" : "tendencia-baja");
            const flecha = Math.abs(s.tendencia) < 5 ? "→" : (s.tendencia > 0 ? "▲" : "▼");
            tendCelda = `<span class="${clase}">${flecha} ${s.tendencia > 0 ? "+" : ""}${s.tendencia.toFixed(0)}%</span>`;
        }
        return `<tr>
            <td>${s.centro}</td>
            <td class="num"><div class="celda-barra"><div class="barra-mini"><div style="width:${anchoBarra}%"></div></div>${fmtNum(s.total)}</div></td>
            <td class="num">${fmtNum(s.promedio)}</td>
            <td class="num">${fmtNum(s.max)}</td>
            <td class="num">${fmtNum(s.min)}</td>
            <td class="num">${participacion.toFixed(1)}%</td>
            <td class="num">${tendCelda}</td>
        </tr>`;
    }).join("");
}

/* ── Comparación mensual de consumo entre establecimientos: una línea por
   centro para comparar directamente su evolución mes a mes (a diferencia del
   apilado de arriba, que muestra composición/total, no la trayectoria de
   cada centro). El color se asigna por identidad del centro (su índice fijo
   en `centros`), no por su ranking, para que no cambie si la selección de
   medicamento/alerta hace que algún centro entre o salga del top. ────────── */
function renderComparacionMensual(){
    const [desde, hasta] = rangoMeses();
    const alertaSet = getAlertaSet();
    const mesesRango = meses.filter(m=>m>=desde && m<=hasta);
    const sub = document.getElementById("comparacionMensualSub");
    const matriz = {}; mesesRango.forEach(m=>matriz[m] = new Array(centros.length).fill(0));
    const totalesPorCentro = new Array(centros.length).fill(0);

    consumo.forEach(row=>{
        if (row.m < desde || row.m > hasta) return;
        if (medicamentoSel !== null) {
            if (row.c !== medicamentoSel) return;
        } else {
            const p = productoByCodigo[row.c];
            if (!p || !alertaSet.has(p.alerta)) return;
        }
        row.v.forEach((val,i)=>{ matriz[row.m][i] += val; totalesPorCentro[i] += val; });
    });

    // Máximo 8 centros a la vez para que el gráfico se pueda leer; se eligen
    // los de mayor consumo en el rango, pero se listan en orden de índice
    // (identidad), no de ranking, así el color de cada uno queda fijo.
    const idxConDatos = totalesPorCentro
        .map((v,i)=>[i,v]).filter(([,v])=>v>0)
        .sort((a,b)=>b[1]-a[1]).slice(0,8)
        .map(([i])=>i).sort((a,b)=>a-b);

    const trazas = idxConDatos.map(i=>({
        x: mesesRango, y: mesesRango.map(m=>matriz[m][i]),
        name: centros[i], mode:"lines+markers", type:"scatter",
        line:{color: PALETA_CENTROS[i % PALETA_CENTROS.length], width:2.5},
        marker:{size:6, color: PALETA_CENTROS[i % PALETA_CENTROS.length]},
        hovertemplate: "%{fullData.name}<br>%{x}<br><b>%{y:,.0f} unid.</b><extra></extra>",
    }));

    sub.textContent = medicamentoSel !== null
        ? `Consumo mensual de "${productoByCodigo[medicamentoSel]?.medicamento ?? medicamentoSel}" por establecimiento (hasta 8 con mayor consumo en el rango).`
        : "Consumo mensual de los productos con la alerta seleccionada, por establecimiento (hasta 8 con mayor consumo en el rango).";

    Plotly.react("chartComparacionMensual", trazas, layoutBase({
        height:440, hovermode:"x unified", margin:{b:70, t:20},
        xaxis:{title:"Mes"}, yaxis:{title:"Unidades consumidas", rangemode:"tozero"},
        legend:{orientation:"h", y:-0.28, font:{size:10}},
    }), PLOTLY_CONFIG);
}

/* ── Proyección de consumo a 3 y 6 meses ─────────────────────────────── */
function renderProy(){
    const activos = productosFiltrados().slice().sort((a,b)=>b.proy6-a.proy6).slice(0,15).reverse();
    const bordeSel = activos.map(p=> p.codigo === medicamentoSel ? "#1F2937" : "rgba(0,0,0,0)");
    const anchoSel  = activos.map(p=> p.codigo === medicamentoSel ? 3 : 0);
    Plotly.react("chartProy", [
        { x: activos.map(p=>p.proy3), y: activos.map(p=>p.medicamento.slice(0,40)),
          name:"Proyección 3 meses", type:"bar", orientation:"h", marker:{color:"#FFD966", line:{color:bordeSel, width:anchoSel}},
          hovertemplate: "%{y}<br><b>%{x:,.0f} unid.</b><extra>3 meses</extra>" },
        { x: activos.map(p=>p.proy6), y: activos.map(p=>p.medicamento.slice(0,40)),
          name:"Proyección 6 meses", type:"bar", orientation:"h", marker:{color:"#1F4E79", line:{color:bordeSel, width:anchoSel}},
          hovertemplate: "%{y}<br><b>%{x:,.0f} unid.</b><extra>6 meses</extra>" },
    ], layoutBase({
        barmode:"group", height:440, margin:{l:10,r:20,t:40,b:40},
        xaxis:{title:"Unidades proyectadas"}, yaxis:{automargin:true},
        legend:{orientation:"h", y:1.1, x:1, xanchor:"right"},
    }), PLOTLY_CONFIG);
}

/* ── Buscador individual: ficha completa del medicamento ─────────────── */
function extraeCodigo(texto){
    const m = /^(\d+)\s*-/.exec(texto.trim());
    return m ? Number(m[1]) : null;
}

function mostrarFicha(codigo){
    const cont = document.getElementById("fichaContenido");
    if (codigo === null){
        cont.innerHTML = '<div class="ficha-vacia">Escribe un código o nombre en el buscador de arriba para ver su ficha completa.</div>';
        renderChartPronostico(null);
        return;
    }
    const p = productoByCodigo[codigo];
    if (!p){ cont.innerHTML = '<div class="ficha-vacia">No se encontró ese medicamento. Elige una opción de la lista.</div>'; return; }

    const badgeClase = "badge-" + p.alerta.replace(/\s+/g,"");
    cont.innerHTML = `
      <div class="ficha-header">
        <div>
          <div class="ficha-titulo">${p.medicamento}</div>
          <div class="ficha-sub">Código ${p.codigo} · ${p.proveedor || "Sin proveedor registrado"}</div>
        </div>
        <span class="badge ${badgeClase}">${p.alerta}</span>
      </div>
      <div class="ficha-kpis">
        <div class="ficha-kpi"><b>${fmtNum(p.saldoBfc)}</b><span>Saldo BFC</span></div>
        <div class="ficha-kpi"><b>${fmtNum(p.minimoBfc)} / ${fmtNum(p.maximoBfc)}</b><span>Mínimo / Máximo BFC</span></div>
        <div class="ficha-kpi"><b>${fmtNum(p.totalSaldos)}</b><span>Total saldos red</span></div>
        <div class="ficha-kpi"><b>${fmtNum(p.consumoMensual)}</b><span>Consumo mensual</span></div>
        <div class="ficha-kpi"><b>${fmtNum(p.proy3)} / ${fmtNum(p.proy6)}</b><span>Proyección 3M / 6M</span></div>
        <div class="ficha-kpi"><b>${fmtNum(p.compra3)} / ${fmtNum(p.compra6)}</b><span>Necesidad compra 3M / 6M</span></div>
        <div class="ficha-kpi"><b title="${fmtCLP(p.valorStockTotal)}">${fmtCompacto(p.valorStockTotal)}</b><span>Valor stock red</span></div>
        <div class="ficha-kpi"><b>${fmtCLP(p.ultPrecio * FACTOR_IVA)}</b><span>Precio unitario de referencia (con IVA)</span></div>
      </div>
      <div class="ficha-graficos">
        <div id="fichaLinea"></div>
        <div id="fichaCentros"></div>
      </div>
      <div id="fichaClasificacion" class="ficha-sub" style="margin-top:10px"></div>`;

    const serie = consumoByCodigo[codigo] || [];
    Plotly.newPlot("fichaLinea", [{
        x: serie.map(r=>r.m), y: serie.map(r=>r.v.reduce((a,b)=>a+b,0)),
        mode:"lines+markers", type:"scatter", line:{color:"#1F4E79", width:2.5}, marker:{size:5},
        fill:"tozeroy", fillcolor:"rgba(31,78,121,0.10)",
        hovertemplate: "%{x}<br><b>%{y:,.0f} unid.</b><extra></extra>",
    }], layoutBase({ title:{text:"Consumo histórico (toda la red)", font:{size:13}}, height:320,
        hovermode:"x unified", margin:{t:36,b:35}, yaxis:{rangemode:"tozero"} }),
    PLOTLY_CONFIG);

    const desglose = new Array(centros.length).fill(0);
    serie.forEach(r=> r.v.forEach((v,i)=> desglose[i]+=v));
    const filas = centros.map((c,i)=>({centro:c, total:desglose[i]}))
        .filter(f=>f.total>0).sort((a,b)=>b.total-a.total).slice(0,12).reverse();

    Plotly.newPlot("fichaCentros", [{
        x: filas.map(f=>f.total), y: filas.map(f=>f.centro),
        type:"bar", orientation:"h", marker:{color:"#70AD47"},
        hovertemplate: "%{y}<br><b>%{x:,.0f} unid.</b><extra></extra>",
    }], layoutBase({ title:{text:"Desglose de consumo por centro (histórico)", font:{size:13}}, height:320,
        margin:{l:10,r:20,t:36,b:35}, yaxis:{automargin:true} }),
    PLOTLY_CONFIG);

    const cls = clasificacion[String(codigo)];
    document.getElementById("fichaClasificacion").textContent = cls
        ? `Clasificación ABC: ${cls.abc} · XYZ: ${cls.xyz}`
        : "Clasificación ABC/XYZ no disponible (sin historia suficiente).";

    renderChartPronostico(codigo);
}

/* ── Selección unificada de medicamento: un solo buscador mueve todos los
   módulos (ficha, línea, desglose por centro, resaltado en ranking/proyección
   y filtro de la tabla) ────────────────────────────────────────────────── */
function seleccionarMedicamento(codigo){
    medicamentoSel = codigo;
    const p = codigo !== null ? productoByCodigo[codigo] : null;
    buscadorInput.value = p ? `${p.codigo} - ${p.medicamento}` : "";
    mostrarFicha(codigo);
    renderTodo();
    renderTodoPronostico();
    renderSaldoEstablecimiento();
}

buscadorInput.addEventListener("input", e=>{
    const texto = e.target.value.trim();
    if (!texto) { seleccionarMedicamento(null); return; }
    const codigo = extraeCodigo(texto);
    if (codigo !== null) seleccionarMedicamento(codigo);
});

/* ── Tabla resumen ordenable y filtrable ─────────────────────────────── */
const COLUMNAS_TABLA = [
    {key:"codigo", label:"Código", num:true},
    {key:"medicamento", label:"Medicamento", num:false},
    {key:"proveedor", label:"Proveedor", num:false},
    {key:"saldoBfc", label:"Saldo BFC", num:true},
    {key:"minimoBfc", label:"Mínimo BFC", num:true},
    {key:"maximoBfc", label:"Máximo BFC", num:true},
    {key:"totalSaldos", label:"Total Saldos", num:true},
    {key:"consumoMensual", label:"Consumo Mensual", num:true},
    {key:"proy3", label:"Proyección 3M", num:true},
    {key:"proy6", label:"Proyección 6M", num:true},
    {key:"compra3", label:"Necesidad Compra 3M", num:true},
    {key:"compra6", label:"Necesidad Compra 6M", num:true},
    {key:"valorStockTotal", label:"Valor Stock Red", num:true},
    {key:"alerta", label:"Alerta", num:false},
];
let sortKey = "compra3", sortDir = -1;

function construyeCabecera(){
    const tr = document.getElementById("tablaHead");
    tr.innerHTML = "";
    COLUMNAS_TABLA.forEach(col=>{
        const th = document.createElement("th");
        th.dataset.key = col.key;
        const flecha = col.key === sortKey ? (sortDir === 1 ? "▲" : "▼") : "";
        th.innerHTML = `${col.label} <span class="arrow">${flecha}</span>`;
        th.addEventListener("click", ()=>{
            if (sortKey === col.key) sortDir *= -1; else { sortKey = col.key; sortDir = col.num ? -1 : 1; }
            construyeCabecera();
            renderTabla();
        });
        tr.appendChild(th);
    });
}

function renderTabla(){
    let filas;
    if (medicamentoSel !== null) {
        // Con un medicamento seleccionado en el buscador global, la tabla se
        // reduce a esa fila (igual criterio que la ficha y el gráfico de línea).
        filas = productos.filter(p=>p.codigo === medicamentoSel);
    } else {
        const alertaSet = getAlertaSet();
        const q = normaliza(document.getElementById("tablaSearch").value.trim());
        filas = productos.filter(p=>alertaSet.has(p.alerta));
        if (q) {
            filas = filas.filter(p =>
                normaliza(p.medicamento).includes(q) ||
                normaliza(p.proveedor).includes(q) ||
                String(p.codigo).includes(q)
            );
        }
    }
    filas = filas.slice().sort((a,b)=>{
        const av = a[sortKey], bv = b[sortKey];
        if (av < bv) return -1 * sortDir;
        if (av > bv) return  1 * sortDir;
        return 0;
    });

    document.getElementById("tablaContador").textContent = `${filas.length} de ${productos.length} productos`;

    const tbody = document.getElementById("tablaBody");
    const frag = document.createDocumentFragment();
    filas.forEach(p=>{
        const tr = document.createElement("tr");
        const badgeClase = "badge-" + p.alerta.replace(/\s+/g,"");
        tr.innerHTML = `
            <td>${p.codigo}</td>
            <td>${p.medicamento}</td>
            <td>${p.proveedor}</td>
            <td class="num">${fmtNum(p.saldoBfc)}</td>
            <td class="num">${fmtNum(p.minimoBfc)}</td>
            <td class="num">${fmtNum(p.maximoBfc)}</td>
            <td class="num">${fmtNum(p.totalSaldos)}</td>
            <td class="num">${fmtNum(p.consumoMensual)}</td>
            <td class="num">${fmtNum(p.proy3)}</td>
            <td class="num">${fmtNum(p.proy6)}</td>
            <td class="num">${fmtNum(p.compra3)}</td>
            <td class="num">${fmtNum(p.compra6)}</td>
            <td class="num" title="${fmtCLP(p.valorStockTotal)}">${fmtCompacto(p.valorStockTotal)}</td>
            <td><span class="badge ${badgeClase}">${p.alerta}</span></td>`;
        frag.appendChild(tr);
    });
    tbody.innerHTML = "";
    tbody.appendChild(frag);
}

/* ── Orquestación de filtros globales ────────────────────────────────── */
function renderTodo(){
    renderKPIs();
    renderLineChart();
    renderRanking();
    renderStack();
    renderConsumoEstablecimientos();
    renderComparacionMensual();
    renderProy();
    renderTabla();
}

[fMesDesde, fMesHasta, fCentro].forEach(el=>el.addEventListener("change", renderTodo));
fAlertaBox.addEventListener("change", renderTodo);
lineCentro.addEventListener("change", renderLineChart);
document.getElementById("tablaSearch").addEventListener("input", renderTabla);

document.getElementById("btnReset").addEventListener("click", ()=>{
    fMesDesde.value = DATA.mesInicioDefecto;
    fMesHasta.value = DATA.mesFinDefecto;
    fCentro.value = "__ALL__";
    fAlertaBox.querySelectorAll("input[type=checkbox]").forEach(chk=>chk.checked = true);
    lineCentro.value = "__ALL__";
    document.getElementById("tablaSearch").value = "";
    seleccionarMedicamento(null);
});

construyeCabecera();
renderTodo();
renderSaldoEstablecimiento();

/* ══════════════════════════════════════════════════════════════════════
   MÓDULO DE PRONÓSTICO DE DEMANDA Y ABASTECIMIENTO
   Reutiliza fmtNum/fmtCLP/fmtCompacto/layoutBase/PLOTLY_CONFIG/normaliza y
   los datos ya cargados (productos, productoByCodigo, consumoByCodigo,
   centros, centroIndex) — no se toca nada de lo anterior.
   ══════════════════════════════════════════════════════════════════════ */
const pronosticos   = DATA.pronosticos   || [];
const clasificacion = DATA.clasificacion || {};
const calidadDatos  = DATA.calidadDatos  || {};
const metodologia   = DATA.metodologia   || {};
const modelosDesc   = metodologia.modelosDesc || {};

const Z_NIVEL_SERVICIO = {"90":1.2816, "95":1.6449, "97.5":1.9600, "99":2.3263};

const pronoPorClave = {};
pronosticos.forEach(pr => { pronoPorClave[pr.c + "|" + pr.e] = pr; });
function pronoRed(codigo){ return pronoPorClave[codigo + "|RED"]; }
function pronoEstablecimiento(codigo, idxCentro){ return pronoPorClave[codigo + "|" + idxCentro]; }

let codigoSeleccionadoPronostico = null;
let filasPronosticoActuales = [];
let pSortKey = "mesesCobertura", pSortDir = 1;

function sumarMeses(mesStr, n){
    const [y, m] = mesStr.split("-").map(Number);
    const d = new Date(y, m - 1 + n, 1);
    return d.getFullYear() + "-" + String(d.getMonth() + 1).padStart(2, "0");
}

function obtenerParametrosPronostico(){
    return {
        horizonte: Number(document.getElementById("pHorizonte").value),
        z: Z_NIVEL_SERVICIO[document.getElementById("pNivelServicio").value] ?? 1.6449,
        ambito: document.getElementById("pAmbitoStock").value,
        leadTimeDias: Math.max(0, Number(document.getElementById("pLeadTime").value) || 0),
        multiplo: Math.max(1, Number(document.getElementById("pMultiplo").value) || 1),
        stockTransito: Math.max(0, Number(document.getElementById("pStockTransito").value) || 0),
        umbralSobrestock: 6,
    };
}

/* ── Indicadores de abastecimiento (se recalculan en vivo, sin volver a
   ejecutar Python, para que cambiar lead time/nivel de servicio/múltiplo
   sea instantáneo). Ámbito de stock: como los archivos Sal_Art_* no traen
   una columna que los vincule con las columnas de establecimiento de
   Consumos_Historicos, el stock solo se puede evaluar a nivel red completa
   o BFC (ver panel de Metodología) ── */
function calcularAbastecimiento(p, pr, params){
    const h = params.horizonte;
    const forecastH = pr.f.slice(0, h);
    const demandaMensualProm = forecastH.reduce((a, b) => a + b, 0) / h;
    const demandaHorizonte = forecastH.reduce((a, b) => a + b, 0);
    const stockDisponible = params.ambito === "bfc" ? p.saldoBfc : p.totalSaldos;
    const leadTimeMeses = params.leadTimeDias / 30;
    const stockSeguridad = params.z * pr.sig * Math.sqrt(Math.max(leadTimeMeses, 1 / 30));

    const mesesCobertura = demandaMensualProm > 0 ? stockDisponible / demandaMensualProm : null;

    let acumulado = 0, mesQuiebreIdx = null;
    for (let i = 0; i < pr.f.length; i++){
        acumulado += pr.f[i];
        if (acumulado >= stockDisponible){ mesQuiebreIdx = i; break; }
    }
    const fechaQuiebreLabel = mesQuiebreIdx === null
        ? "Sin quiebre en 12 meses" : sumarMeses(DATA.mesFinDefecto, mesQuiebreIdx + 1);

    const bruto = demandaHorizonte + stockSeguridad - stockDisponible - params.stockTransito;
    const cantidadSugerida = bruto > 0 ? Math.ceil(bruto / params.multiplo) * params.multiplo : 0;
    const gastoProyectado = cantidadSugerida * p.ultPrecio;
    const riesgoSobrestock = mesesCobertura !== null && mesesCobertura > params.umbralSobrestock;

    let semaforo;
    if (demandaMensualProm <= 0){
        semaforo = stockDisponible > 0 ? "AZUL" : "VERDE";
    } else if (mesQuiebreIdx !== null && mesQuiebreIdx <= Math.ceil(leadTimeMeses)){
        semaforo = "ROJO";
    } else if (mesesCobertura < leadTimeMeses + 1){
        semaforo = "AMARILLO";
    } else if (riesgoSobrestock){
        semaforo = "AZUL";
    } else {
        semaforo = "VERDE";
    }

    return { demandaMensualProm, demandaHorizonte, stockDisponible, stockSeguridad,
        mesesCobertura, fechaQuiebreLabel, cantidadSugerida, gastoProyectado, riesgoSobrestock, semaforo };
}

function construirFilaPronostico(p, pr, params){
    const cls = clasificacion[String(p.codigo)] || { abc: "C", xyz: "N/D" };
    const ab = calcularAbastecimiento(p, pr, params);
    return {
        codigo: p.codigo, medicamento: p.medicamento, proveedor: p.proveedor,
        abc: cls.abc, xyz: cls.xyz, modelo: pr.mod, modeloDesc: modelosDesc[pr.mod] || pr.mod,
        nivelConfianza: pr.nc, wape: pr.wape,
        mesesCobertura: ab.mesesCobertura, fechaQuiebreLabel: ab.fechaQuiebreLabel,
        cantidadSugerida: ab.cantidadSugerida, gastoProyectado: ab.gastoProyectado,
        demandaHorizonte: ab.demandaHorizonte, semaforo: ab.semaforo,
    };
}

function construirFilasPronostico(){
    const params = obtenerParametrosPronostico();

    if (medicamentoSel !== null){
        // Con un medicamento seleccionado en el buscador global, la tabla de
        // priorización se reduce a esa fila (igual criterio que la tabla resumen).
        const p = productoByCodigo[medicamentoSel];
        const pr = p ? pronoRed(medicamentoSel) : null;
        return (p && pr) ? [construirFilaPronostico(p, pr, params)] : [];
    }

    const metodoSel = document.getElementById("pMetodo").value;
    const abcSel = document.getElementById("pABC").value;
    const xyzSel = document.getElementById("pXYZ").value;
    const catSel = document.getElementById("pCategoria").value;
    const q = normaliza(document.getElementById("pTablaSearch").value.trim());

    const filas = [];
    productos.forEach(p => {
        const pr = pronoRed(p.codigo);
        if (!pr) return;
        const cls = clasificacion[String(p.codigo)] || { abc: "C", xyz: "N/D" };
        if (abcSel !== "__ALL__" && cls.abc !== abcSel) return;
        if (xyzSel !== "__ALL__" && cls.xyz !== xyzSel) return;
        if (metodoSel !== "__ALL__" && pr.mod !== metodoSel) return;
        if (catSel !== "__ALL__" && (p.categoria || "Sin categoría") !== catSel) return;
        if (q){
            const hit = normaliza(p.medicamento).includes(q) || normaliza(p.proveedor).includes(q) || String(p.codigo).includes(q);
            if (!hit) return;
        }
        filas.push(construirFilaPronostico(p, pr, params));
    });
    return filas;
}

/* ── KPIs ─────────────────────────────────────────────────────────────── */
function renderKPIsPronostico(filas){
    const cont = document.getElementById("kpiPronostico");
    const consumoHist = filas.reduce((s, f) => {
        const p = productoByCodigo[f.codigo]; return s + (p ? p.consumoMensual * 12 : 0);
    }, 0);
    const demandaProy = filas.reduce((s, f) => s + f.demandaHorizonte, 0);
    const gastoProy = filas.reduce((s, f) => s + f.gastoProyectado, 0);
    const coberturas = filas.map(f => f.mesesCobertura).filter(v => v != null && isFinite(v));
    const coberturaProm = coberturas.length ? coberturas.reduce((a, b) => a + b, 0) / coberturas.length : null;
    const nRojo = filas.filter(f => f.semaforo === "ROJO").length;

    cont.innerHTML = `
      <div class="kpi-card" style="border-top-color:var(--header)">
        <div class="kpi-valor" style="color:var(--header)">${fmtNum(consumoHist)}</div>
        <div class="kpi-etiqueta">CONSUMO ÚLTIMOS 12M</div>
      </div>
      <div class="kpi-card" style="border-top-color:var(--header)">
        <div class="kpi-valor" style="color:var(--header)">${fmtNum(demandaProy)}</div>
        <div class="kpi-etiqueta">DEMANDA PROYECTADA</div>
      </div>
      <div class="kpi-card" style="border-top-color:var(--header)" title="${fmtCLP(gastoProy)}">
        <div class="kpi-valor" style="color:var(--header)">${fmtCompacto(gastoProy)}</div>
        <div class="kpi-etiqueta">GASTO PROYECTADO</div>
      </div>
      <div class="kpi-card" style="border-top-color:var(--ok)">
        <div class="kpi-valor" style="color:var(--ok)">${coberturaProm != null ? coberturaProm.toFixed(1) : "s/d"}</div>
        <div class="kpi-etiqueta">COBERTURA PROMEDIO (MESES)</div>
      </div>
      <div class="kpi-card" style="border-top-color:var(--urgente)">
        <div class="kpi-valor" style="color:var(--urgente)">${fmtNum(nRojo)}</div>
        <div class="kpi-etiqueta">PRODUCTOS EN RIESGO DE QUIEBRE</div>
      </div>`;
}

/* ── Gráfico de consumo + pronóstico + intervalos (producto seleccionado) ── */
function renderChartPronostico(codigo){
    codigoSeleccionadoPronostico = codigo;
    const info = document.getElementById("infoModeloSerie");
    if (codigo === null){
        Plotly.purge("chartPronostico");
        info.textContent = "Selecciona un medicamento en el buscador de arriba para ver su pronóstico.";
        return;
    }
    const p = productoByCodigo[codigo];

    const centroSel = fCentro.value;
    let pr = null, etiqueta = "toda la red";
    if (centroSel !== "__ALL__" && centroIndex[centroSel] !== undefined){
        const candidato = pronoEstablecimiento(codigo, centroIndex[centroSel]);
        if (candidato){ pr = candidato; etiqueta = centroSel; }
    }
    if (!pr) pr = pronoRed(codigo);

    if (!p || !pr){
        Plotly.purge("chartPronostico");
        info.textContent = "No hay pronóstico disponible para este producto/establecimiento (historia insuficiente o sin consumo real).";
        return;
    }

    const params = obtenerParametrosPronostico();
    const h = params.horizonte;
    const serieHist = consumoByCodigo[codigo] || [];
    const xHist = serieHist.map(r => r.m);
    const yHist = serieHist.map(r => r.v.reduce((a, b) => a + b, 0));

    const mesesFuturos = [];
    for (let i = 1; i <= h; i++) mesesFuturos.push(sumarMeses(DATA.mesFinDefecto, i));

    const trazas = [
        { x: xHist, y: yHist, name: "Consumo histórico", mode: "lines+markers", type: "scatter",
          line: { color: "#1F4E79", width: 2.5 }, marker: { size: 5 },
          hovertemplate: "%{x}<br><b>%{y:,.0f} unid.</b><extra>Histórico</extra>" },
        { x: mesesFuturos, y: pr.u9.slice(0, h), name: "Límite superior 95%", mode: "lines",
          line: { width: 0 }, showlegend: false, hoverinfo: "skip" },
        { x: mesesFuturos, y: pr.l9.slice(0, h), name: "Intervalo 95%", mode: "lines",
          line: { width: 0 }, fill: "tonexty", fillcolor: "rgba(31,78,121,0.10)", hoverinfo: "skip" },
        { x: mesesFuturos, y: pr.u8.slice(0, h), name: "Límite superior 80%", mode: "lines",
          line: { width: 0 }, showlegend: false, hoverinfo: "skip" },
        { x: mesesFuturos, y: pr.l8.slice(0, h), name: "Intervalo 80%", mode: "lines",
          line: { width: 0 }, fill: "tonexty", fillcolor: "rgba(31,78,121,0.22)", hoverinfo: "skip" },
        { x: mesesFuturos, y: pr.f.slice(0, h), name: "Pronóstico", mode: "lines+markers", type: "scatter",
          line: { color: "#C0504D", width: 2.5, dash: "dash" }, marker: { size: 5 },
          hovertemplate: "%{x}<br><b>%{y:,.0f} unid.</b><extra>Pronóstico</extra>" },
    ];

    Plotly.react("chartPronostico", trazas, layoutBase({
        title: { text: `${p.medicamento} — consumo e intervalos (${etiqueta})`, font: { size: 14 } },
        height: 420, hovermode: "x unified", margin: { t: 44 },
        xaxis: { title: "Mes" }, yaxis: { title: "Unidades", rangemode: "tozero" },
        legend: { orientation: "h", y: -0.22, font: { size: 10 } },
    }), PLOTLY_CONFIG);

    const wapeTxt = pr.wape != null ? (pr.wape * 100).toFixed(1) + "%" : "sin datos suficientes";
    info.innerHTML = `Modelo seleccionado: <b>${modelosDesc[pr.mod] || pr.mod}</b> · `
        + `Confianza: <b>${pr.nc}</b> · WAPE (backtest): <b>${wapeTxt}</b> · `
        + `Sesgo: <b>${pr.sesgo != null ? fmtNum(pr.sesgo) : "s/d"}</b>`;
}

/* ── Tabla de priorización por riesgo ─────────────────────────────────── */
const COLUMNAS_PRONOSTICO = [
    { key: "codigo", label: "Código", num: true },
    { key: "medicamento", label: "Medicamento", num: false },
    { key: "abc", label: "ABC", num: false },
    { key: "xyz", label: "XYZ", num: false },
    { key: "modeloDesc", label: "Modelo", num: false },
    { key: "nivelConfianza", label: "Confianza", num: false },
    { key: "wape", label: "WAPE", num: true },
    { key: "mesesCobertura", label: "Cobertura (meses)", num: true },
    { key: "fechaQuiebreLabel", label: "Quiebre estimado", num: false },
    { key: "cantidadSugerida", label: "Compra sugerida", num: true },
    { key: "gastoProyectado", label: "Gasto proyectado", num: true },
    { key: "semaforo", label: "Semáforo", num: false },
];

function construyeCabeceraPronostico(){
    const tr = document.getElementById("pTablaHead");
    tr.innerHTML = "";
    COLUMNAS_PRONOSTICO.forEach(col => {
        const th = document.createElement("th");
        const flecha = col.key === pSortKey ? (pSortDir === 1 ? "▲" : "▼") : "";
        th.innerHTML = `${col.label} <span class="arrow">${flecha}</span>`;
        th.addEventListener("click", () => {
            if (pSortKey === col.key) pSortDir *= -1; else { pSortKey = col.key; pSortDir = 1; }
            construyeCabeceraPronostico();
            renderTablaDesdeFilas(filasPronosticoActuales);
        });
        tr.appendChild(th);
    });
}

function renderTablaDesdeFilas(filas){
    const ordenadas = filas.slice().sort((a, b) => {
        let av = a[pSortKey], bv = b[pSortKey];
        if (av === null || av === undefined) av = pSortDir === 1 ? Infinity : -Infinity;
        if (bv === null || bv === undefined) bv = pSortDir === 1 ? Infinity : -Infinity;
        if (av < bv) return -1 * pSortDir;
        if (av > bv) return 1 * pSortDir;
        return 0;
    });

    document.getElementById("pTablaContador").textContent = `${ordenadas.length} de ${productos.length} productos con pronóstico`;

    const tbody = document.getElementById("pTablaBody");
    const frag = document.createDocumentFragment();
    ordenadas.forEach(f => {
        const tr = document.createElement("tr");
        tr.innerHTML = `
            <td>${f.codigo}</td>
            <td>${f.medicamento}</td>
            <td>${f.abc}</td>
            <td>${f.xyz}</td>
            <td>${f.modeloDesc}</td>
            <td>${f.nivelConfianza}</td>
            <td class="num">${f.wape != null ? (f.wape * 100).toFixed(1) + "%" : "s/d"}</td>
            <td class="num">${f.mesesCobertura != null ? f.mesesCobertura.toFixed(1) : "s/c"}</td>
            <td>${f.fechaQuiebreLabel}</td>
            <td class="num">${fmtNum(f.cantidadSugerida)}</td>
            <td class="num" title="${fmtCLP(f.gastoProyectado)}">${fmtCompacto(f.gastoProyectado)}</td>
            <td><span class="badge badge-${f.semaforo}">${f.semaforo}</span></td>`;
        frag.appendChild(tr);
    });
    tbody.innerHTML = "";
    tbody.appendChild(frag);
}

function renderTodoPronostico(){
    const filas = construirFilasPronostico();
    filasPronosticoActuales = filas;
    renderKPIsPronostico(filas);
    renderTablaDesdeFilas(filas);
    if (codigoSeleccionadoPronostico !== null) renderChartPronostico(codigoSeleccionadoPronostico);
    renderGeneradorPedido();
}

/* ── Descarga CSV / Excel (sin librerías externas, 100% offline) ────────── */
function csvEsc(s){
    s = String(s ?? "");
    return /[;"\n]/.test(s) ? '"' + s.replace(/"/g, '""') + '"' : s;
}
function descargarBlob(contenido, nombre, tipo){
    const blob = new Blob(["﻿" + contenido], { type: tipo });
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url; a.download = nombre;
    document.body.appendChild(a); a.click(); document.body.removeChild(a);
    URL.revokeObjectURL(url);
}
const ENCABEZADOS_EXPORT = ["Código", "Medicamento", "Proveedor", "ABC", "XYZ", "Modelo", "Confianza",
    "WAPE", "Cobertura (meses)", "Quiebre estimado", "Compra sugerida", "Gasto proyectado", "Semáforo"];
function filaExport(f){
    return [f.codigo, f.medicamento, f.proveedor, f.abc, f.xyz, f.modeloDesc, f.nivelConfianza,
        f.wape != null ? (f.wape * 100).toFixed(1) + "%" : "", f.mesesCobertura != null ? f.mesesCobertura.toFixed(1) : "",
        f.fechaQuiebreLabel, f.cantidadSugerida, Math.round(f.gastoProyectado), f.semaforo];
}
function exportarCSV(){
    const lineas = [ENCABEZADOS_EXPORT.join(";")];
    filasPronosticoActuales.forEach(f => lineas.push(filaExport(f).map(csvEsc).join(";")));
    descargarBlob(lineas.join("\n"), "pronostico_abastecimiento.csv", "text/csv;charset=utf-8;");
}
function exportarExcel(){
    let html = "<table><tr>" + ENCABEZADOS_EXPORT.map(h => `<th>${h}</th>`).join("") + "</tr>";
    filasPronosticoActuales.forEach(f => {
        html += "<tr>" + filaExport(f).map(v => `<td>${String(v ?? "").replace(/</g, "&lt;")}</td>`).join("") + "</tr>";
    });
    html += "</table>";
    descargarBlob(html, "pronostico_abastecimiento.xls", "application/vnd.ms-excel");
}

/* ── Solicitud de compra (usa como modelo el formulario SOLICITUD: DETALLE DE
   REQUERIMIENTO DE BIENES Y/O SERVICIOS UGR) ─────────────────────────────── */
const FACTOR_IVA = 1.19; // el precio del archivo de origen viene sin IVA

function construirFilasSolicitud(horizonteMeses){
    const params = Object.assign({}, obtenerParametrosPronostico(), { horizonte: horizonteMeses });
    const filas = [];
    productos.forEach(p => {
        const pr = pronoRed(p.codigo);
        if (!pr) return;
        const ab = calcularAbastecimiento(p, pr, params);
        const enRiesgo = ab.semaforo === "ROJO" || ab.semaforo === "AMARILLO";
        if (ab.cantidadSugerida > 0 && enRiesgo){
            const precioConIva = p.ultPrecio * FACTOR_IVA;
            filas.push({
                codigo: p.codigo, medicamento: p.medicamento, proveedor: p.proveedor,
                cantidad: ab.cantidadSugerida, precioUnitario: precioConIva,
                subtotal: ab.cantidadSugerida * precioConIva, semaforo: ab.semaforo,
            });
        }
    });
    const ORDEN_RIESGO = { ROJO: 0, AMARILLO: 1 };
    filas.sort((a, b) => (ORDEN_RIESGO[a.semaforo] - ORDEN_RIESGO[b.semaforo]) || a.medicamento.localeCompare(b.medicamento));
    return filas;
}

function exportarSolicitud(){
    const horizonte = Number(document.getElementById("solHorizonte").value);
    const filas = construirFilasSolicitud(horizonte);
    const cont = document.getElementById("solContador");
    if (!filas.length){
        cont.textContent = "Ningún producto en riesgo ROJO/AMARILLO necesita compra en ese horizonte con los parámetros actuales.";
        return;
    }
    const hoy = new Date();
    const fechaLabel = [hoy.getDate(), hoy.getMonth() + 1, hoy.getFullYear()]
        .map((n, i) => i < 2 ? String(n).padStart(2, "0") : n).join("-");
    const totalGeneral = filas.reduce((s, f) => s + f.subtotal, 0);
    const esc = s => String(s ?? "").replace(/</g, "&lt;");

    let html = `<table>
        <tr><td colspan="12" style="font-size:15px;font-weight:bold;">DETALLE DE REQUERIMIENTO DE BIENES Y/O SERVICIOS UGR</td></tr>
        <tr><td colspan="12"></td></tr>
        <tr><td><b>Unidad solicitante</b></td><td colspan="3">Droguería BFC</td><td><b>Fecha</b></td><td colspan="2">${fechaLabel}</td></tr>
        <tr><td><b>Horizonte de la solicitud</b></td><td colspan="3">${horizonte} meses</td><td><b>Ámbito de stock</b></td><td colspan="2">${document.getElementById("pAmbitoStock").selectedOptions[0].text}</td></tr>
        <tr><td colspan="12">Incluye solo productos en riesgo ROJO (quiebre probable) o AMARILLO (cobertura insuficiente).</td></tr>
        <tr><td colspan="12"></td></tr>
        <tr>
            <th>ITEM</th><th>CANTIDAD</th><th>UNIDAD DE MEDIDA</th><th>ID CONVENIO MARCO</th>
            <th>CODIGO AVIS</th><th>PRODUCTO O SERVICIO</th><th>DETALLE (marca, modelo, color, etc.)</th>
            <th>IMAGEN DEL PRODUCTO</th><th>ENLACE A SITIO WEB DONDE SE VENDA EL PRODUCTO SOLICITADO</th>
            <th>PRECIO UNITARIO DE REFERENCIA (con IVA)</th><th>SUBTOTAL (con IVA)</th><th>RIESGO</th>
        </tr>`;
    filas.forEach((f, i) => {
        html += `<tr>
            <td>${i + 1}</td><td>${fmtNum(f.cantidad)}</td><td></td><td></td>
            <td>${f.codigo}</td><td>${esc(f.medicamento)}</td><td>${esc(f.proveedor)}</td>
            <td></td><td></td>
            <td>${fmtCLP(f.precioUnitario)}</td><td>${fmtCLP(f.subtotal)}</td><td>${f.semaforo}</td>
        </tr>`;
    });
    html += `<tr><td colspan="9"></td><td><b>TOTAL</b></td><td><b>${fmtCLP(totalGeneral)}</b></td><td></td></tr>
    </table>`;

    cont.textContent = `${filas.length} productos en riesgo ROJO/AMARILLO a solicitar · Monto total ${fmtCLP(totalGeneral)}`;
    descargarBlob(html, `Solicitud_Compra_${horizonte}M_${fechaLabel}.xls`, "application/vnd.ms-excel");
}

/* ── Generador de Pedido con horizonte variable ────────────────────────────
   Complementa la "Solicitud de compra" (que usa tramos fijos de 3/6/12 meses
   y el semáforo de riesgo): aquí el horizonte se elige libremente en días y
   la lista se arma con el criterio literal que pide el negocio — "productos
   que tendrán quiebre proyectado en ese horizonte de tiempo" — interpolando
   el pronóstico mensual para estimar el día en que el stock disponible se
   agota. ──────────────────────────────────────────────────────────────── */
const DIAS_POR_MES = 30.4375; // promedio calendario, para convertir días <-> meses de pronóstico

// Demanda acumulada del pronóstico hasta un horizonte fraccionario de meses
// (p.ej. 3.5 meses = 3 meses completos + la mitad del 4to mes de pronóstico).
function demandaHastaHorizonteMeses(pr, horizonteMeses){
    const capMeses = Math.min(Math.max(horizonteMeses, 0), pr.f.length);
    const mesesCompletos = Math.floor(capMeses);
    let total = 0;
    for (let i = 0; i < mesesCompletos; i++) total += pr.f[i];
    const resto = capMeses - mesesCompletos;
    if (resto > 0 && mesesCompletos < pr.f.length) total += pr.f[mesesCompletos] * resto;
    return total;
}

// Días hasta que la demanda acumulada del pronóstico supere el stock
// disponible, interpolando linealmente dentro del mes en que ocurre el
// cruce. Devuelve null si no se proyecta quiebre dentro de los 12 meses de
// pronóstico disponibles.
function diasHastaQuiebre(pr, stockDisponible){
    if (stockDisponible <= 0) return 0;
    let acumulado = 0;
    for (let i = 0; i < pr.f.length; i++){
        const previo = acumulado;
        acumulado += pr.f[i];
        if (acumulado >= stockDisponible){
            const demandaMes = pr.f[i];
            const fraccion = demandaMes > 0 ? (stockDisponible - previo) / demandaMes : 0;
            return (i + Math.max(0, Math.min(1, fraccion))) * DIAS_POR_MES;
        }
    }
    return null;
}

function fechaDesdeDiasForecast(dias){
    const [y, m] = DATA.mesFinDefecto.split("-").map(Number);
    const base = new Date(y, m, 1); // primer día del mes siguiente al último dato histórico
    const fecha = new Date(base.getTime() + dias * 86400000);
    return fecha.toLocaleDateString("es-CL", { day: "2-digit", month: "2-digit", year: "numeric" });
}

function calcularPedidoHorizonte(p, pr, horizonteDias, paramsBase){
    const horizonteMeses = horizonteDias / DIAS_POR_MES;
    const stockDisponible = paramsBase.ambito === "bfc" ? p.saldoBfc : p.totalSaldos;
    const diasQuiebre = diasHastaQuiebre(pr, stockDisponible);
    const dentroHorizonte = diasQuiebre !== null && diasQuiebre <= horizonteDias;
    const demandaHorizonte = demandaHastaHorizonteMeses(pr, horizonteMeses);
    const leadTimeMeses = paramsBase.leadTimeDias / 30;
    const stockSeguridad = paramsBase.z * pr.sig * Math.sqrt(Math.max(leadTimeMeses, 1 / 30));
    const bruto = demandaHorizonte + stockSeguridad - stockDisponible - paramsBase.stockTransito;
    const cantidadSugerida = bruto > 0 ? Math.ceil(bruto / paramsBase.multiplo) * paramsBase.multiplo : 0;
    const precioConIva = p.ultPrecio * FACTOR_IVA;
    return {
        dentroHorizonte, diasQuiebre,
        fechaQuiebreLabel: diasQuiebre !== null ? fechaDesdeDiasForecast(diasQuiebre) : "Sin quiebre en 12 meses",
        stockDisponible, demandaHorizonte, cantidadSugerida,
        costo: cantidadSugerida * precioConIva,
    };
}

function obtenerHorizonteDias(){
    return Number(document.getElementById("pedHorizonteDias").value) || 90;
}

function actualizarLabelHorizonte(){
    const dias = obtenerHorizonteDias();
    const mesesEq = dias / DIAS_POR_MES;
    document.getElementById("pedHorizonteLabel").textContent = `${dias} días (~${mesesEq.toFixed(1)} meses)`;
}

function construirFilasPedido(){
    const paramsBase = obtenerParametrosPronostico();
    const horizonteDias = obtenerHorizonteDias();
    const q = normaliza(document.getElementById("pedTablaSearch").value.trim());

    // Con un medicamento seleccionado en el buscador global, la lista se
    // reduce a ese producto (mismo criterio que el resto de las secciones).
    const universo = medicamentoSel !== null
        ? productos.filter(p => p.codigo === medicamentoSel)
        : productos;

    const filas = [];
    universo.forEach(p => {
        const pr = pronoRed(p.codigo);
        if (!pr) return;
        if (q){
            const hit = normaliza(p.medicamento).includes(q) || normaliza(p.proveedor).includes(q) || String(p.codigo).includes(q);
            if (!hit) return;
        }
        const ped = calcularPedidoHorizonte(p, pr, horizonteDias, paramsBase);
        if (!ped.dentroHorizonte) return; // solo productos con quiebre proyectado dentro del horizonte elegido
        filas.push(Object.assign(
            { codigo: p.codigo, medicamento: p.medicamento, proveedor: p.proveedor, consumoMensual: p.consumoMensual },
            ped
        ));
    });
    filas.sort((a, b) => a.diasQuiebre - b.diasQuiebre);
    return filas;
}

function renderKPIsGeneradorPedido(filas){
    const cont = document.getElementById("kpiGeneradorPedido");
    const totalUnidades = filas.reduce((s, f) => s + f.cantidadSugerida, 0);
    const totalCosto = filas.reduce((s, f) => s + f.costo, 0);
    const masProximo = filas.length ? Math.round(filas[0].diasQuiebre) : null;
    cont.innerHTML = `
      <div class="kpi-card" style="border-top-color:var(--urgente)">
        <div class="kpi-valor" style="color:var(--urgente)">${fmtNum(filas.length)}</div>
        <div class="kpi-etiqueta">PRODUCTOS CON QUIEBRE EN EL HORIZONTE</div>
      </div>
      <div class="kpi-card" style="border-top-color:var(--header)">
        <div class="kpi-valor" style="color:var(--header)">${fmtNum(totalUnidades)}</div>
        <div class="kpi-etiqueta">UNIDADES A PEDIR</div>
      </div>
      <div class="kpi-card" style="border-top-color:var(--header)" title="${fmtCLP(totalCosto)}">
        <div class="kpi-valor" style="color:var(--header)">${fmtCompacto(totalCosto)}</div>
        <div class="kpi-etiqueta">COSTO ESTIMADO DEL PEDIDO</div>
      </div>
      <div class="kpi-card" style="border-top-color:var(--normal)">
        <div class="kpi-valor" style="color:#5c4a00">${masProximo !== null ? fmtNum(masProximo) : "—"}</div>
        <div class="kpi-etiqueta">QUIEBRE MÁS PRÓXIMO (DÍAS)</div>
      </div>`;
}

const COLUMNAS_PEDIDO = [
    { key: "codigo", label: "Código", num: true },
    { key: "medicamento", label: "Medicamento", num: false },
    { key: "proveedor", label: "Proveedor", num: false },
    { key: "stockDisponible", label: "Stock disponible", num: true },
    { key: "consumoMensual", label: "Consumo mensual", num: true },
    { key: "diasQuiebre", label: "Días hasta quiebre", num: true },
    { key: "fechaQuiebreLabel", label: "Fecha estimada de quiebre", num: false },
    { key: "cantidadSugerida", label: "Cantidad a pedir", num: true },
    { key: "costo", label: "Costo estimado", num: true },
];
let pedSortKey = "diasQuiebre", pedSortDir = 1;
let filasPedidoActuales = [];

function construyeCabeceraPedido(){
    const tr = document.getElementById("pedTablaHead");
    tr.innerHTML = "";
    COLUMNAS_PEDIDO.forEach(col => {
        const th = document.createElement("th");
        const flecha = col.key === pedSortKey ? (pedSortDir === 1 ? "▲" : "▼") : "";
        th.innerHTML = `${col.label} <span class="arrow">${flecha}</span>`;
        th.addEventListener("click", () => {
            if (pedSortKey === col.key) pedSortDir *= -1; else { pedSortKey = col.key; pedSortDir = 1; }
            construyeCabeceraPedido();
            renderTablaPedido(filasPedidoActuales);
        });
        tr.appendChild(th);
    });
}

function renderTablaPedido(filas){
    const ordenadas = filas.slice().sort((a, b) => {
        let av = a[pedSortKey], bv = b[pedSortKey];
        if (av === null || av === undefined) av = pedSortDir === 1 ? Infinity : -Infinity;
        if (bv === null || bv === undefined) bv = pedSortDir === 1 ? Infinity : -Infinity;
        if (av < bv) return -1 * pedSortDir;
        if (av > bv) return 1 * pedSortDir;
        return 0;
    });

    document.getElementById("pedTablaContador").textContent =
        `${ordenadas.length} productos con quiebre proyectado dentro del horizonte elegido`;

    const tbody = document.getElementById("pedTablaBody");
    const frag = document.createDocumentFragment();
    ordenadas.forEach(f => {
        const tr = document.createElement("tr");
        tr.innerHTML = `
            <td>${f.codigo}</td>
            <td>${f.medicamento}</td>
            <td>${f.proveedor}</td>
            <td class="num">${fmtNum(f.stockDisponible)}</td>
            <td class="num">${fmtNum(f.consumoMensual)}</td>
            <td class="num">${Math.round(f.diasQuiebre)}</td>
            <td>${f.fechaQuiebreLabel}</td>
            <td class="num">${fmtNum(f.cantidadSugerida)}</td>
            <td class="num" title="${fmtCLP(f.costo)}">${fmtCompacto(f.costo)}</td>`;
        frag.appendChild(tr);
    });
    tbody.innerHTML = "";
    tbody.appendChild(frag);
}

function renderGeneradorPedido(){
    actualizarLabelHorizonte();
    const filas = construirFilasPedido();
    filasPedidoActuales = filas;
    renderKPIsGeneradorPedido(filas);
    renderTablaPedido(filas);
}

function exportarPedidoHorizonte(){
    const filas = filasPedidoActuales;
    if (!filas.length){
        document.getElementById("pedTablaContador").textContent =
            "Ningún producto tiene quiebre proyectado dentro del horizonte elegido con los parámetros actuales.";
        return;
    }
    const horizonteDias = obtenerHorizonteDias();
    const hoy = new Date();
    const fechaLabel = [hoy.getDate(), hoy.getMonth() + 1, hoy.getFullYear()]
        .map((n, i) => i < 2 ? String(n).padStart(2, "0") : n).join("-");
    const totalGeneral = filas.reduce((s, f) => s + f.costo, 0);
    const esc = s => String(s ?? "").replace(/</g, "&lt;");

    let html = `<table>
        <tr><td colspan="12" style="font-size:15px;font-weight:bold;">DETALLE DE REQUERIMIENTO DE BIENES Y/O SERVICIOS UGR</td></tr>
        <tr><td colspan="12"></td></tr>
        <tr><td><b>Unidad solicitante</b></td><td colspan="3">Droguería BFC</td><td><b>Fecha</b></td><td colspan="2">${fechaLabel}</td></tr>
        <tr><td><b>Horizonte del pedido</b></td><td colspan="3">${horizonteDias} días (~${(horizonteDias / DIAS_POR_MES).toFixed(1)} meses)</td><td><b>Ámbito de stock</b></td><td colspan="2">${document.getElementById("pAmbitoStock").selectedOptions[0].text}</td></tr>
        <tr><td colspan="12">Incluye solo productos con quiebre de stock proyectado dentro de ese horizonte, según el pronóstico de demanda.</td></tr>
        <tr><td colspan="12"></td></tr>
        <tr>
            <th>ITEM</th><th>CANTIDAD</th><th>UNIDAD DE MEDIDA</th><th>ID CONVENIO MARCO</th>
            <th>CODIGO AVIS</th><th>PRODUCTO O SERVICIO</th><th>DETALLE (marca, modelo, color, etc.)</th>
            <th>IMAGEN DEL PRODUCTO</th><th>ENLACE A SITIO WEB DONDE SE VENDA EL PRODUCTO SOLICITADO</th>
            <th>PRECIO UNITARIO DE REFERENCIA (con IVA)</th><th>SUBTOTAL (con IVA)</th><th>DÍAS HASTA QUIEBRE</th>
        </tr>`;
    filas.forEach((f, i) => {
        const precioUnitario = f.cantidadSugerida > 0 ? f.costo / f.cantidadSugerida : 0;
        html += `<tr>
            <td>${i + 1}</td><td>${fmtNum(f.cantidadSugerida)}</td><td></td><td></td>
            <td>${f.codigo}</td><td>${esc(f.medicamento)}</td><td>${esc(f.proveedor)}</td>
            <td></td><td></td>
            <td>${fmtCLP(precioUnitario)}</td><td>${fmtCLP(f.costo)}</td><td>${Math.round(f.diasQuiebre)}</td>
        </tr>`;
    });
    html += `<tr><td colspan="9"></td><td><b>TOTAL</b></td><td><b>${fmtCLP(totalGeneral)}</b></td><td></td></tr>
    </table>`;

    descargarBlob(html, `Pedido_Quiebre_${horizonteDias}d_${fechaLabel}.xls`, "application/vnd.ms-excel");
}

/* ── Metodología (estática, se arma una sola vez) ────────────────────────── */
function renderMetodologia(){
    const m = metodologia, cal = calidadDatos;
    document.getElementById("metodologiaDatos").innerHTML = `
      <div class="metodologia-dato"><b>${m.periodoEntrenamiento || "s/d"}</b><span>Período de entrenamiento</span></div>
      <div class="metodologia-dato"><b>${(m.horizontes || []).join(" / ")} meses</b><span>Horizontes de pronóstico</span></div>
      <div class="metodologia-dato"><b>${DATA.generadoEl}</b><span>Fecha de última actualización</span></div>
      <div class="metodologia-dato"><b>${fmtNum(cal.n_series_con_pronostico)}</b><span>Series código+establecimiento pronosticadas</span></div>`;
    document.getElementById("metodologiaReglas").innerHTML = (m.reglasHistoria || []).map(t => `<li>${t}</li>`).join("");
    document.getElementById("metodologiaMetricas").textContent = (m.metricas || []).join(", ");
    document.getElementById("metodologiaSupuestos").innerHTML = (m.supuestos || []).map(t => `<li>${t}</li>`).join("");
    document.getElementById("metodologiaLimitaciones").innerHTML = (m.limitaciones || []).map(t => `<li>${t}</li>`).join("");
    document.getElementById("metodologiaCalidad").innerHTML = `
      <li>${fmtNum(cal.n_codigos)} códigos analizados, ${fmtNum(cal.codigos_historia_insuficiente)} con historia insuficiente (menos de 12 meses).</li>
      <li>${fmtNum(cal.duplicados_codigo_fecha)} filas código+fecha duplicadas detectadas y consolidadas.</li>
      <li>${fmtNum(cal.valores_negativos_corregidos)} valores negativos detectados y corregidos a 0.</li>
      <li>${fmtNum(cal.outliers_detectados)} valores atípicos detectados (no eliminados, solo señalados).</li>
      <li>${fmtNum(cal.total_meses_ausentes)} meses sin registro detectados en total (distinto de consumo cero explícito).</li>`;
}

function poblarFiltrosPronostico(){
    const modelosPresentes = new Set();
    pronosticos.forEach(pr => { if (pr.e === "RED") modelosPresentes.add(pr.mod); });
    const pMetodo = document.getElementById("pMetodo");
    Array.from(modelosPresentes).sort().forEach(mo => pMetodo.appendChild(new Option(modelosDesc[mo] || mo, mo)));

    const categorias = new Set(productos.map(p => p.categoria || "Sin categoría"));
    const pCategoria = document.getElementById("pCategoria");
    Array.from(categorias).sort().forEach(c => pCategoria.appendChild(new Option(c, c)));
}

["pHorizonte", "pNivelServicio", "pAmbitoStock", "pLeadTime", "pMultiplo", "pStockTransito",
 "pMetodo", "pABC", "pXYZ", "pCategoria"].forEach(id => {
    const el = document.getElementById(id);
    el.addEventListener("change", renderTodoPronostico);
    if (el.tagName === "INPUT") el.addEventListener("input", renderTodoPronostico);
});
document.getElementById("pTablaSearch").addEventListener("input", renderTodoPronostico);
fCentro.addEventListener("change", () => {
    if (codigoSeleccionadoPronostico !== null) renderChartPronostico(codigoSeleccionadoPronostico);
});
document.getElementById("btnResetPronostico").addEventListener("click", () => {
    document.getElementById("pHorizonte").value = "12";
    document.getElementById("pNivelServicio").value = "95";
    document.getElementById("pAmbitoStock").value = "red";
    document.getElementById("pLeadTime").value = 15;
    document.getElementById("pMultiplo").value = 1;
    document.getElementById("pStockTransito").value = 0;
    document.getElementById("pMetodo").value = "__ALL__";
    document.getElementById("pABC").value = "__ALL__";
    document.getElementById("pXYZ").value = "__ALL__";
    document.getElementById("pCategoria").value = "__ALL__";
    document.getElementById("pTablaSearch").value = "";
    document.getElementById("pedHorizonteDias").value = 90;
    document.getElementById("pedTablaSearch").value = "";
    renderTodoPronostico();
});
document.getElementById("btnDescargarCSV").addEventListener("click", exportarCSV);
document.getElementById("btnDescargarExcel").addEventListener("click", exportarExcel);
document.getElementById("btnDescargarSolicitud").addEventListener("click", exportarSolicitud);
document.getElementById("pedHorizonteDias").addEventListener("input", renderGeneradorPedido);
document.getElementById("pedTablaSearch").addEventListener("input", renderGeneradorPedido);
document.getElementById("btnDescargarPedido").addEventListener("click", exportarPedidoHorizonte);

construyeCabeceraPronostico();
construyeCabeceraPedido();
poblarFiltrosPronostico();
renderMetodologia();
renderChartPronostico(null);
renderTodoPronostico();
</script>
</body>
</html>
"""

html = (PLANTILLA
        .replace("__PLOTLYJS__", plotly_js)
        .replace("__DATA_JSON__", data_json))

RUTA_DASHBOARD.write_text(html, encoding="utf-8")
print(f"Tamaño del archivo: {RUTA_DASHBOARD.stat().st_size / 1024 / 1024:.1f} MB")

print("\n" + "=" * 70)
print(f"✔ DASHBOARD ACTUALIZADO: {RUTA_DASHBOARD}")
print(f"  Generado el {datetime.now().strftime('%d-%m-%Y %H:%M')}")
print("=" * 70)
