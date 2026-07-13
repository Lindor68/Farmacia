"""
Genera resumen_stock_BFC.xlsx a partir de:
  - consolidado.xlsx           -> stock consolidado por bodega (generado por generar_pedido.py)
  - Consumos_Historicos.xlsx   -> consumo mensual histórico por bodega

Resume saldos y necesidades de compra centradas en la bodega de droguería (BFC).
Basta con volver a ejecutar generar_pedido.py (que regenera consolidado.xlsx) y luego
este script cada vez que se actualicen los archivos de origen.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

carpeta = Path(__file__).resolve().parent

BODEGA_BFC             = "Sal_Art_BFC"
BODEGAS_EXCLUIR_STOCK   = {"Sal_Art_ETHON"}   # no son puntos reales de stock de la red
PATRON_BODEGA_TRANSITO  = "ransversal"        # "Unidad Técnica Transversal" (UTTO): tránsito interno, no consumo real
MESES_PROYECCION        = (3, 6)

# ════════════════════════════════════════════════════════════════════════════
# 1. STOCK — leer consolidado.xlsx
# ════════════════════════════════════════════════════════════════════════════
cons = pd.read_excel(carpeta / "consolidado.xlsx")
cons["Código"]       = pd.to_numeric(cons["Código"], errors="coerce")
cons["Stock Actual"] = pd.to_numeric(cons["Stock Actual"], errors="coerce").fillna(0)
cons["Stock Mínimo"] = pd.to_numeric(cons["Stock Mínimo"], errors="coerce").fillna(0)
cons["Stock Máx."]   = pd.to_numeric(cons["Stock Máx."], errors="coerce").fillna(0)
cons = cons.dropna(subset=["Código"])
cons["Código"] = cons["Código"].astype(int)

bfc = cons[cons["origen"] == BODEGA_BFC].copy()
if bfc.empty:
    raise ValueError(f"No se encontró la bodega '{BODEGA_BFC}' en consolidado.xlsx")

bfc["Ult. Precio Ingresado"] = pd.to_numeric(bfc["Ult. Precio Ingresado"], errors="coerce")

saldo_bfc = bfc.groupby("Código", as_index=False).agg(
    **{"Saldo BFC":  ("Stock Actual", "sum"),
       "Mínimo BFC": ("Stock Mínimo", "sum"),
       "Máximo BFC": ("Stock Máx.",   "sum"),
       "Ult. Precio": ("Ult. Precio Ingresado", "max")}
)

red = cons[~cons["origen"].isin(BODEGAS_EXCLUIR_STOCK)]
total_saldos = (red.groupby("Código", as_index=False)["Stock Actual"].sum()
                 .rename(columns={"Stock Actual": "Total Saldos"}))

resto = red[red["origen"] != BODEGA_BFC]
total_max_resto = (resto.groupby("Código", as_index=False)["Stock Máx."].sum()
                    .rename(columns={"Stock Máx.": "Total Máximos (excepto BFC)"}))

nombres = (cons.drop_duplicates("Código")[["Código", "Nombre Artículo", "Proveedor"]]
           .rename(columns={"Nombre Artículo": "Medicamento"}))

print(f"Stock: {cons['Código'].nunique()} productos en consolidado.xlsx")

# ════════════════════════════════════════════════════════════════════════════
# 2. CONSUMO — leer Consumos_Historicos.xlsx
#    (el archivo trae los nombres de "Código"/"Artículo" con la codificación
#    dañada de origen; se toman por posición para no depender de esos bytes)
# ════════════════════════════════════════════════════════════════════════════
ch = pd.read_excel(carpeta / "Consumos_Historicos.xlsx", sheet_name="Hoja1")
col_codigo, col_articulo = ch.columns[0], ch.columns[1]
ch = ch.rename(columns={col_codigo: "Codigo", col_articulo: "Articulo"})
ch["Codigo"] = pd.to_numeric(ch["Codigo"], errors="coerce")
ch["Fecha"]  = pd.to_datetime(ch["Fecha"])
ch = ch.dropna(subset=["Codigo"])
ch["Codigo"] = ch["Codigo"].astype(int)

cols_bodega  = [c for c in ch.columns if c not in ("Codigo", "Articulo", "TOTAL", "Fecha")]
col_transito = [c for c in cols_bodega if PATRON_BODEGA_TRANSITO in c and "Sur" not in c]
cols_consumo = [c for c in cols_bodega if c not in col_transito]

# Consumo mensual promedio de los últimos 12 meses con datos (dinámico según
# la fecha más reciente del archivo, igual criterio que generar_pedido.py)
fecha_max    = ch["Fecha"].max()
fecha_inicio = fecha_max - pd.DateOffset(months=11)
reciente = ch[ch["Fecha"] >= fecha_inicio].copy()
reciente[cols_consumo] = reciente[cols_consumo].apply(pd.to_numeric, errors="coerce").fillna(0)
reciente["consumo_red"] = reciente[cols_consumo].sum(axis=1)

meses_por_codigo = reciente.groupby("Codigo")["Fecha"].nunique()
consumo = reciente.groupby("Codigo", as_index=False)["consumo_red"].sum()
consumo["meses"]           = consumo["Codigo"].map(meses_por_codigo).fillna(1)
consumo["Consumo Mensual"] = (consumo["consumo_red"] / consumo["meses"]).round(2)
consumo = consumo.rename(columns={"Codigo": "Código"})[["Código", "Consumo Mensual"]]

arsenal = (ch[["Codigo", "Articulo"]].drop_duplicates("Codigo")
           .rename(columns={"Codigo": "Código"}))

print(f"Consumo calculado sobre {fecha_inicio.date()} a {fecha_max.date()} "
      f"({len(arsenal)} productos, bodega de tránsito excluida: {col_transito})")

# ════════════════════════════════════════════════════════════════════════════
# 3. CRUCE
# ════════════════════════════════════════════════════════════════════════════
df = arsenal.merge(saldo_bfc,       on="Código", how="left")
df = df.merge(total_saldos,         on="Código", how="left")
df = df.merge(total_max_resto,      on="Código", how="left")
df = df.merge(nombres,              on="Código", how="left")
df = df.merge(consumo,              on="Código", how="left")

for col in ["Saldo BFC", "Mínimo BFC", "Máximo BFC", "Ult. Precio", "Total Saldos",
            "Total Máximos (excepto BFC)", "Consumo Mensual"]:
    df[col] = df[col].fillna(0)
df["Proveedor"]   = df["Proveedor"].fillna("")
df["Medicamento"] = df["Medicamento"].fillna(df["Articulo"])

# Valorización de stock al último precio ingresado en BFC (mismo criterio que
# generar_resumen_stock_compra.py, del que depende generar_dashboard.py)
df["Valor Stock BFC"]   = (df["Saldo BFC"]    * df["Ult. Precio"]).round(0).astype(int)
df["Valor Stock Total"] = (df["Total Saldos"] * df["Ult. Precio"]).round(0).astype(int)

# ════════════════════════════════════════════════════════════════════════════
# 4. PROYECCIONES, NECESIDAD DE COMPRA Y ALERTAS
# ════════════════════════════════════════════════════════════════════════════
for m in MESES_PROYECCION:
    df[f"Proyección Consumo {m}M"] = (df["Consumo Mensual"] * m).round(0).astype(int)
    df[f"Necesidad Compra {m}M"]   = (
        (df[f"Proyección Consumo {m}M"] - df["Total Saldos"]).clip(lower=0).round(0).astype(int)
    )

# Alerta según meses de cobertura de la red completa (Total Saldos / Consumo Mensual)
consumo_nz      = df["Consumo Mensual"].replace(0, np.nan)
meses_cobertura = df["Total Saldos"] / consumo_nz

def alerta(m):
    if pd.isna(m): return "SIN CONSUMO"
    if m < 1:      return "URGENTE"
    if m <= 2:     return "NORMAL"
    return "OK"

df["Alerta Stock"] = meses_cobertura.map(alerta)

orden_a = {"URGENTE": 0, "NORMAL": 1, "OK": 2, "SIN CONSUMO": 3}
df["_ord"] = df["Alerta Stock"].map(orden_a)
df = df.sort_values(["_ord", "Necesidad Compra 3M"], ascending=[True, False]).drop(columns="_ord")

cols_salida = [
    "Código", "Medicamento", "Proveedor",
    "Saldo BFC", "Mínimo BFC", "Máximo BFC",
    "Total Saldos", "Total Máximos (excepto BFC)",
    "Ult. Precio", "Valor Stock BFC", "Valor Stock Total",
    "Consumo Mensual",
    "Proyección Consumo 3M", "Proyección Consumo 6M",
    "Necesidad Compra 3M", "Necesidad Compra 6M",
    "Alerta Stock",
]
salida = df[cols_salida].reset_index(drop=True)

print(f"\nProductos en resumen: {len(salida)}")
print(f"Alertas:\n{salida['Alerta Stock'].value_counts().to_string()}")

# ════════════════════════════════════════════════════════════════════════════
# 5. GUARDAR CON FORMATO
# ════════════════════════════════════════════════════════════════════════════
fill_urg  = PatternFill(fill_type="solid", fgColor="FF4444")
fill_nor  = PatternFill(fill_type="solid", fgColor="FFD966")
fill_ok   = PatternFill(fill_type="solid", fgColor="70AD47")
fill_sc   = PatternFill(fill_type="solid", fgColor="BFBFBF")
fill_hdr  = PatternFill(fill_type="solid", fgColor="1F4E79")
font_hdr  = Font(bold=True, color="FFFFFF", size=10)
font_bold = Font(bold=True)
align_c   = Alignment(horizontal="center", vertical="center")

anchos = [10, 50, 28, 14, 14, 14, 16, 24, 14, 18, 18, 18, 20, 20, 18, 18, 14]

def escribir_excel(path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        salida.to_excel(writer, index=False, sheet_name="Resumen Stock BFC")
        ws = writer.sheets["Resumen Stock BFC"]

        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        for cell in ws[1]:
            cell.fill      = fill_hdr
            cell.font      = font_hdr
            cell.alignment = align_c
        ws.row_dimensions[1].height = 30

        colores = {"URGENTE": fill_urg, "NORMAL": fill_nor, "OK": fill_ok, "SIN CONSUMO": fill_sc}
        col_alerta = len(cols_salida)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            alerta_val = row[col_alerta - 1].value
            fill       = colores.get(alerta_val)
            if fill:
                row[col_alerta - 1].fill = fill
                row[col_alerta - 1].font = font_bold
            for cell in row[2:]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row[1].alignment = Alignment(horizontal="left", vertical="center")

        ws.freeze_panes = "A2"


out_primary = carpeta / "resumen_stock_BFC.xlsx"
out_backup  = carpeta / "resumen_stock_BFC_backup.xlsx"

try:
    escribir_excel(out_primary)
    print(f"\nArchivo guardado: {out_primary}")
except PermissionError:
    escribir_excel(out_backup)
    print(f"\n{out_primary.name} está abierto en Excel y no se pudo sobrescribir.")
    print(f"Se guardó como: {out_backup}")
    print("Ciérralo y vuelve a ejecutar el script, o renombra el archivo de respaldo manualmente.")
