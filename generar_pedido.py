"""
Genera pedido_proyectado.xlsx a partir de:
  - Sal_Art_*.xlsx   -> stock actual por bodega (exportados desde el sistema)
  - Consumos_historicos.csv -> arsenal de droguería + consumo mensual histórico

Basta con reemplazar esos archivos en esta carpeta y volver a ejecutar el script;
no hace falta tocar el código ni generar archivos intermedios a mano.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

carpeta = Path(__file__).resolve().parent

BODEGA_DROGUERIA   = "Sal_Art_BFC"     # bodega que gestiona droguería (compra/distribuye)
BODEGAS_EXCLUIR_STOCK_TOTAL = {"Sal_Art_ETHON"}  # no son puntos reales de stock de la red
BODEGAS_EXCLUIR_CONSUMO     = {"UTTO"}           # columna de tránsito interno, no consumo real

# ════════════════════════════════════════════════════════════════════════════
# 1. LEER STOCK: consolidar todos los Sal_Art_*.xlsx de la carpeta
# ════════════════════════════════════════════════════════════════════════════
def leer_archivo_stock(ruta):
    """Cada export tiene 2 filas de cabecera (con celdas combinadas); las detecta
    automáticamente en vez de asumir un número de fila fijo."""
    wb = load_workbook(ruta, data_only=True)
    ws = wb.active

    fila_cab = None
    for i, fila in enumerate(ws.iter_rows(values_only=True), start=1):
        no_vacios = sum(1 for c in fila if c is not None and str(c).strip() != "")
        if no_vacios >= 5:
            fila_cab = i
            break
    if fila_cab is None:
        return pd.DataFrame()

    cab1 = [c.value for c in ws[fila_cab]]
    cab2 = [c.value for c in ws[fila_cab + 1]]

    columnas = []
    for a, b in zip(cab1, cab2):
        a = str(a).strip() if a is not None else ""
        b = str(b).strip() if b is not None else ""
        if a and b:
            columnas.append(f"{a} {b}")
        elif a:
            columnas.append(a)
        elif b:
            columnas.append(b)
        else:
            columnas.append("Sin nombre")

    filas = []
    for fila in ws.iter_rows(min_row=fila_cab + 2, values_only=True):
        if any(c is not None and str(c).strip() != "" for c in fila):
            filas.append(list(fila))

    df = pd.DataFrame(filas, columns=columnas)
    df = df.loc[:, ~df.columns.str.startswith("Sin nombre")]
    df = df.dropna(how="all")
    # En algunas bodegas "Código de Barras" viene vacío (dtype object) y en otras
    # trae valores numéricos (dtype float64); se fija como texto en todas para
    # que el dtype sea consistente antes de concatenar los archivos.
    if "Código de Barras" in df.columns:
        df["Código de Barras"] = df["Código de Barras"].astype("string")
    return df


archivos_stock = sorted(carpeta.glob("Sal_Art_*.xlsx"))
if not archivos_stock:
    raise FileNotFoundError(f"No se encontraron archivos Sal_Art_*.xlsx en {carpeta}")

trozos = []
for archivo in archivos_stock:
    origen = archivo.stem
    df = leer_archivo_stock(archivo)
    if df.empty:
        print(f"  Aviso: {archivo.name} no aportó filas, se omite")
        continue
    df.insert(0, "origen", origen)
    trozos.append(df)

cons = pd.concat(trozos, ignore_index=True, sort=False)
cons["Código"]       = pd.to_numeric(cons["Código"], errors="coerce")
cons["Stock Actual"] = pd.to_numeric(cons["Stock Actual"], errors="coerce").fillna(0)
print(f"Stock consolidado: {len(cons)} filas de {len(archivos_stock)} bodegas")

# Guardamos el consolidado para poder auditar el cruce si algo no cuadra
cons.to_excel(carpeta / "consolidado.xlsx", index=False)

# ════════════════════════════════════════════════════════════════════════════
# 2. ARSENAL Y CONSUMO: productos que gestiona droguería = los del CSV histórico
# ════════════════════════════════════════════════════════════════════════════
consumo_df = pd.read_csv(
    carpeta / "Consumos_historicos.csv",
    encoding="latin-1", sep=None, engine="python",
)
# "Código"/"Artículo" llegan con la codificación dañada de origen del sistema
# que exporta el archivo; se identifican por posición en vez de por nombre literal
# (mismo criterio que generar_resumen_stock_compra.py).
col_codigo, col_articulo = consumo_df.columns[0], consumo_df.columns[1]
consumo_df = consumo_df.rename(columns={col_codigo: "Codigo", col_articulo: "Producto"})
consumo_df["Producto"] = consumo_df["Producto"].str.lstrip("﻿\xff").str.strip()
consumo_df["Fecha"]    = pd.to_datetime(consumo_df["Fecha"], dayfirst=True)
consumo_df["Codigo"]   = pd.to_numeric(consumo_df["Codigo"], errors="coerce")

arsenal = consumo_df[["Codigo", "Producto"]].drop_duplicates("Codigo").dropna(subset=["Codigo"])
arsenal["Codigo"] = arsenal["Codigo"].astype(int)
print(f"Arsenal droguería: {len(arsenal)} productos")

# Consumo mensual promedio de los últimos 12 meses con datos (dinámico: se ajusta
# solo a la fecha más reciente del CSV cada vez que se actualiza)
fecha_max    = consumo_df["Fecha"].max()
fecha_inicio = fecha_max - pd.DateOffset(months=11)

cols_id      = {"Codigo", "Producto", "TOTAL", "Fecha"}
cols_bodegas = [c for c in consumo_df.columns if c not in cols_id and c not in BODEGAS_EXCLUIR_CONSUMO]

reciente = consumo_df[consumo_df["Fecha"] >= fecha_inicio].copy()
reciente[cols_bodegas] = reciente[cols_bodegas].apply(pd.to_numeric, errors="coerce").fillna(0)
reciente["consumo_red"] = reciente[cols_bodegas].sum(axis=1)

meses_por_producto = reciente.groupby("Codigo")["Fecha"].nunique()
consumo_agg = reciente.groupby("Codigo", as_index=False)["consumo_red"].sum()
consumo_agg["meses"]           = consumo_agg["Codigo"].map(meses_por_producto).fillna(1)
consumo_agg["Consumo Mensual"] = (consumo_agg["consumo_red"] / consumo_agg["meses"]).round(2)
consumo_agg["Codigo"]          = consumo_agg["Codigo"].astype(int)
print(f"Consumo calculado sobre {fecha_inicio.date()} a {fecha_max.date()}")

# ════════════════════════════════════════════════════════════════════════════
# 3. STOCK POR BODEGA
# ════════════════════════════════════════════════════════════════════════════
bfc = cons[cons["origen"] == BODEGA_DROGUERIA].copy()
if bfc.empty:
    raise ValueError(f"No se encontró la bodega de droguería '{BODEGA_DROGUERIA}' en los archivos Sal_Art_*.xlsx")

bfc["Stock Mínimo"]          = pd.to_numeric(bfc["Stock Mínimo"], errors="coerce").fillna(0)
bfc["Stock Máx."]            = pd.to_numeric(bfc["Stock Máx."], errors="coerce").fillna(0)
bfc["Ult. Precio Ingresado"] = pd.to_numeric(bfc["Ult. Precio Ingresado"], errors="coerce")

stock_drog = bfc.groupby("Código", as_index=False).agg(
    **{"Stock Droguería":    ("Stock Actual",           "sum"),
       "Mínimo Droguería":   ("Stock Mínimo",           "sum"),
       "Máximo Droguería":   ("Stock Máx.",             "sum"),
       "Ult. Precio":        ("Ult. Precio Ingresado",  "max"),
       "Proveedor":          ("Proveedor",              "first")}
)
stock_drog.rename(columns={"Código": "Codigo"}, inplace=True)
stock_drog["Codigo"] = stock_drog["Codigo"].astype(int)

total = cons[~cons["origen"].isin(BODEGAS_EXCLUIR_STOCK_TOTAL)]
stock_total = total.groupby("Código", as_index=False)["Stock Actual"].sum()
stock_total.rename(columns={"Código": "Codigo", "Stock Actual": "Stock Total"}, inplace=True)
stock_total["Codigo"] = stock_total["Codigo"].astype(int)

nombres = (
    total.drop_duplicates(subset="Código")[["Código", "Nombre Artículo"]]
    .rename(columns={"Código": "Codigo", "Nombre Artículo": "Medicamento"})
)
nombres["Codigo"] = nombres["Codigo"].astype(int)

# ════════════════════════════════════════════════════════════════════════════
# 4. CRUCE — solo productos del arsenal de droguería
# ════════════════════════════════════════════════════════════════════════════
df = arsenal.merge(consumo_agg[["Codigo", "Consumo Mensual"]], on="Codigo", how="left")
df = df.merge(stock_drog,  on="Codigo", how="left")
df = df.merge(stock_total, on="Codigo", how="left")
df = df.merge(nombres,     on="Codigo", how="left")

for col, relleno in [
    ("Consumo Mensual",  0), ("Stock Droguería", 0), ("Mínimo Droguería", 0),
    ("Máximo Droguería", 0), ("Ult. Precio",     0), ("Stock Total",      0),
]:
    df[col] = df[col].fillna(relleno)
df["Proveedor"] = df["Proveedor"].fillna("")

# Nombre canónico: el del stock si existe, si no el del histórico de consumo
df["Medicamento"] = df["Medicamento"].fillna(df["Producto"])

# ════════════════════════════════════════════════════════════════════════════
# 5. CÁLCULOS
# ════════════════════════════════════════════════════════════════════════════
df["Proyección 3 Meses"] = (df["Consumo Mensual"] * 3 - df["Stock Total"]).clip(lower=0).round(0).astype(int)
df["Proyección 6 Meses"] = (df["Consumo Mensual"] * 6 - df["Stock Total"]).clip(lower=0).round(0).astype(int)

# Compra Necesaria: lo que debe comprar droguería para tener 3 meses en su bodega
df["Compra Necesaria"] = (df["Consumo Mensual"] * 3 - df["Stock Droguería"]).clip(lower=0).round(0).astype(int)

# Prioridad según meses cubiertos por el stock de toda la red
consumo_nz = df["Consumo Mensual"].replace(0, np.nan)
meses_cub  = df["Stock Total"] / consumo_nz

def prioridad(m):
    if pd.isna(m): return "SIN CONSUMO"
    if m < 1:      return "URGENTE"
    if m <= 2:     return "NORMAL"
    return "OK"

df["Prioridad"] = meses_cub.map(prioridad)

# ════════════════════════════════════════════════════════════════════════════
# 6. ORDENAR Y SELECCIONAR COLUMNAS FINALES
# ════════════════════════════════════════════════════════════════════════════
orden_p = {"URGENTE": 0, "NORMAL": 1, "OK": 2, "SIN CONSUMO": 3}
df["_ord"] = df["Prioridad"].map(orden_p)
df = df.sort_values(["_ord", "Compra Necesaria"], ascending=[True, False]).drop(columns="_ord")

cols_salida = [
    "Medicamento", "Proveedor",
    "Stock Droguería", "Mínimo Droguería", "Máximo Droguería", "Ult. Precio",
    "Stock Total", "Consumo Mensual",
    "Proyección 3 Meses", "Proyección 6 Meses", "Compra Necesaria", "Prioridad",
]
salida = df[cols_salida].reset_index(drop=True)

print(f"\nProductos en pedido proyectado: {len(salida)}")
print(f"Prioridades:\n{salida['Prioridad'].value_counts().to_string()}")

# ════════════════════════════════════════════════════════════════════════════
# 7. GUARDAR CON FORMATO
# ════════════════════════════════════════════════════════════════════════════
fill_urg  = PatternFill(fill_type="solid", fgColor="FF4444")
fill_nor  = PatternFill(fill_type="solid", fgColor="FFD966")
fill_ok   = PatternFill(fill_type="solid", fgColor="70AD47")
fill_sc   = PatternFill(fill_type="solid", fgColor="BFBFBF")
fill_hdr  = PatternFill(fill_type="solid", fgColor="1F4E79")
font_hdr  = Font(bold=True, color="FFFFFF", size=10)
font_bold = Font(bold=True)
align_c   = Alignment(horizontal="center", vertical="center")

anchos = [55, 30, 18, 18, 18, 16, 14, 26, 22, 22, 20, 14]

def escribir_excel(path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        salida.to_excel(writer, index=False, sheet_name="Pedido Proyectado")
        ws = writer.sheets["Pedido Proyectado"]

        for i, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        for cell in ws[1]:
            cell.fill      = fill_hdr
            cell.font      = font_hdr
            cell.alignment = align_c
        ws.row_dimensions[1].height = 30

        colores = {"URGENTE": fill_urg, "NORMAL": fill_nor, "OK": fill_ok, "SIN CONSUMO": fill_sc}
        col_prio = len(cols_salida)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            prio_val = row[col_prio - 1].value
            fill     = colores.get(prio_val)
            if fill:
                row[col_prio - 1].fill = fill
                row[col_prio - 1].font = font_bold
            for cell in row[1:]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row[0].alignment = Alignment(horizontal="left", vertical="center")

        ws.freeze_panes = "A2"


out_primary = carpeta / "pedido_proyectado.xlsx"
out_backup  = carpeta / "pedido_proyectado_backup.xlsx"

try:
    escribir_excel(out_primary)
    print(f"\nArchivo guardado: {out_primary}")
except PermissionError:
    escribir_excel(out_backup)
    print(f"\n{out_primary.name} está abierto en Excel y no se pudo sobrescribir.")
    print(f"Se guardó como: {out_backup}")
    print("Ciérralo y vuelve a ejecutar el script, o renombra el archivo de respaldo manualmente.")
